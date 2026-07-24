#!/usr/bin/env python3
"""千葉県公式Excel名簿を学校検索用JSONへ変換する。

原本は data-source/chiba/2025/ に置く（Git管理外）。

- 公立(5-1-r7_koritsu_meibo.xls, .xls形式=xlrdで読む):
  幼稚園/認定こども園/小学校/中学校/義務教育学校/中等教育学校/
  高等学校(全日制・定時制・通信制の3シート)/特別支援学校 の10シート構成。
- 補正(r8-gakkouichiran-2_henkou_ichiran.xls): 令和7→8年度の
  新設・廃止（統合）・所在地変更・名称変更・幼稚園休廃園一覧（千葉市除く）。
- 私立4ファイル(.xlsx=openpyxlで読む): 幼稚園/小学校/中学校(中等前期を含む)/
  高等学校(中等後期を含む)。

学校種の扱い（ユーザー確定方針）:
  - 幼保連携型認定こども園は school_type を独立表記のまま保持する（幼稚園に統合しない）。
  - 高等学校は課程が複数あっても1校1レコードとし、course に配列で保持する。
    所在地・電話番号・学校名のいずれかが課程間で不一致の場合のみ自動統合せず、
    警告を出して別レコードのまま残す。
  - 私立の中等教育学校は前期課程・後期課程を1レコードに統合し、
    school_type を「中等教育学校」に統一する。学校名・住所等が一致しない場合は
    統合せず警告を出す。

出力レコードのフィールド（ユーザー確定スキーマ + 実装上の補助フィールド）:
  id, prefecture, name, name_kana, postal_code, address, municipality,
  school_type, establishment, operator, phone, website, source_name,
  source_url, source_date, verified_date, course
"""

from __future__ import annotations

import argparse
import datetime
import json
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 正規化ユーティリティ
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\n", "").replace("\r", "")
    return text.strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"[ \t　]+", "", text)


def compact_header(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"[\s　・･()（）\[\]【】]", "", text)
    return text.lower()


def normalize_postal_code(value: Any) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    text = normalize_text(value).replace("〒", "")
    return text


PHONE_NUMBER_RE = re.compile(r"0\d{1,4}-\d{1,4}-\d{3,4}")


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    text = normalize_text(value).replace("−", "-").replace("―", "-").replace("ー", "-")

    # 原本で「（全）04-xxxx-xxxx （通）04-xxxx-xxxx」のように課程別に複数の電話番号が
    # 併記されている場合、課程ラベルの丸括弧は落とし、番号のみを " / " で連結して保持する。
    # 番号自体は原本記載値をそのまま使う（推測・合成はしない）。
    numbers = PHONE_NUMBER_RE.findall(text)
    if len(numbers) >= 2:
        return " / ".join(numbers)

    if "-" in text:
        return text
    digits = re.sub(r"\D", "", text)
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return text


def normalize_address(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("−", "-").replace("―", "-")
    text = re.sub(r"[ \t　]+", "", text)
    if text and not text.startswith("千葉県"):
        text = "千葉県" + text
    return text


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).lower()
    normalized = re.sub(r"[^0-9a-zぁ-んァ-ヶ一-龠]+", "-", normalized).strip("-")
    return normalized or "school"


def excel_serial_to_iso(value: Any) -> str:
    """xlrdのExcelシリアル値(float)または既存datetimeを ISO 日付文字列へ変換する。"""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return normalize_text(value)
    if serial <= 0:
        return normalize_text(value)
    epoch = datetime.date(1899, 12, 30)
    return (epoch + datetime.timedelta(days=int(serial))).isoformat()


# ---------------------------------------------------------------------------
# 市町村表示順（千葉市各区 -> 市部 -> 郡部）
# ユーザー確定方針: 千葉市は区ごとに分割し、住所から区名が確定できる場合は
# 地区列より住所を優先する。
# ---------------------------------------------------------------------------

CHIBA_CITY_WARDS = [
    "千葉市中央区", "千葉市花見川区", "千葉市稲毛区",
    "千葉市若葉区", "千葉市緑区", "千葉市美浜区",
]

# 市部（東葛 -> 印旛 -> 香取 -> 海匝 -> 京葉 -> 山武 -> 長生 -> 夷隅 -> 君津 -> 安房、
# 千葉県教育委員会「市町村一覧(エリア別)」の掲載順に準拠）
CHIBA_CITIES = [
    # 東葛
    "我孫子市", "柏市", "鎌ケ谷市", "流山市", "野田市", "松戸市",
    # 印旛
    "印西市", "佐倉市", "白井市", "富里市", "成田市", "八街市", "四街道市",
    # 香取
    "香取市",
    # 海匝
    "旭市", "匝瑳市", "銚子市",
    # 京葉
    "市川市", "市原市", "浦安市", "習志野市", "船橋市", "八千代市",
    # 山武
    "大網白里市", "山武市", "東金市",
    # 長生
    "茂原市",
    # 夷隅
    "いすみ市", "勝浦市",
    # 君津
    "木更津市", "君津市", "袖ケ浦市", "富津市",
    # 安房
    "鴨川市", "館山市", "南房総市",
]

# 郡部（印旛郡 -> 香取郡 -> 山武郡 -> 長生郡 -> 夷隅郡 -> 安房郡）
CHIBA_GUN_TOWNS = [
    "印旛郡酒々井町", "印旛郡栄町",
    "香取郡神崎町", "香取郡多古町", "香取郡東庄町",
    "山武郡九十九里町", "山武郡芝山町", "山武郡横芝光町",
    "長生郡一宮町", "長生郡睦沢町", "長生郡長生村",
    "長生郡白子町", "長生郡長柄町", "長生郡長南町",
    "夷隅郡大多喜町", "夷隅郡御宿町",
    "安房郡鋸南町",
]

MUNICIPALITY_ORDER = CHIBA_CITY_WARDS + CHIBA_CITIES + CHIBA_GUN_TOWNS

# 郡名を省略した表記（原本の住所列に郡名が付かない場合の救済）
_BARE_TOWN_TO_CANONICAL = {
    town[len(gun):]: town
    for gun, town in (
        (m.group(1), m.group(0))
        for m in (re.match(r"(印旛郡|香取郡|山武郡|長生郡|夷隅郡|安房郡)(.+)", t) for t in CHIBA_GUN_TOWNS)
    )
}

_MUNICIPALITY_CANDIDATES = sorted(MUNICIPALITY_ORDER, key=len, reverse=True)
_BARE_TOWN_CANDIDATES = sorted(_BARE_TOWN_TO_CANONICAL, key=len, reverse=True)


def infer_municipality(address: str) -> str:
    text = address
    if text.startswith("千葉県"):
        text = text[len("千葉県"):]
    for candidate in _MUNICIPALITY_CANDIDATES:
        if text.startswith(candidate):
            return candidate
    for bare in _BARE_TOWN_CANDIDATES:
        if text.startswith(bare):
            return _BARE_TOWN_TO_CANONICAL[bare]
    return ""


SCHOOL_TYPE_ORDER = [
    "幼稚園", "幼保連携型認定こども園", "小学校", "中学校",
    "義務教育学校", "高等学校", "中等教育学校", "特別支援学校",
]

HIGH_SCHOOL_COURSE_ORDER = ["全日制", "定時制", "通信制"]


@dataclass
class Warning_:
    context: str
    message: str


WARNINGS: list[Warning_] = []


def warn(context: str, message: str) -> None:
    WARNINGS.append(Warning_(context=context, message=message))


# ---------------------------------------------------------------------------
# レコード生成の共通ヘルパー
# ---------------------------------------------------------------------------

def make_record(
    *,
    name: str,
    name_kana: str,
    postal_code: str,
    address: str,
    school_type: str,
    establishment: str,
    operator: str,
    phone: str,
    website: str,
    source_name: str,
    source_url: str,
    source_date: str,
    course: list[str],
) -> dict[str, Any]:
    name = normalize_name(name)
    address = normalize_address(address)
    municipality = infer_municipality(address)
    postal_code = normalize_postal_code(postal_code)
    phone = normalize_phone(phone)
    stable_key = "|".join((establishment, school_type, municipality, name, ",".join(course)))
    return {
        "id": f"chiba-{slug(stable_key)}",
        "prefecture": "千葉県",
        "name": name,
        "name_kana": normalize_name(name_kana),
        "postal_code": postal_code,
        "address": address,
        "municipality": municipality,
        "school_type": school_type,
        "establishment": establishment,
        "operator": normalize_name(operator),
        "phone": phone,
        "website": website,
        "source_name": source_name,
        "source_url": source_url,
        "source_date": source_date,
        "verified_date": "",
        "course": list(course),
    }


# ---------------------------------------------------------------------------
# 公立学校名簿（5-1-r7_koritsu_meibo.xls, xlrd）
# ---------------------------------------------------------------------------

PUBLIC_HEADER_ALIASES = {
    "committee": ("委員会名",),
    "name": ("学校名", "幼稚園名", "こども園名", "園名"),
    "postal": ("〒", "郵便", "郵便番号"),
    "address": ("所在地",),
    "phone": ("tel", "電話番号", "電話"),
}

PUBLIC_SHEETS = [
    # (sheet_name, school_type, course_label)
    ("5-1-1(幼稚園)", "幼稚園", ""),
    ("5-1-2(認定こども園)", "幼保連携型認定こども園", ""),
    ("5-1-3(小学校)", "小学校", ""),
    ("5-1-4(中学校)", "中学校", ""),
    ("5-1-5(義務教育学校)", "義務教育学校", ""),
    ("5-1-6(中等教育学校)", "中等教育学校", ""),
    ("5-1-(7-1)(高等学校(全日制))", "高等学校", "全日制"),
    ("5-1-(7-2)(高等学校(定時制))", "高等学校", "定時制"),
    ("5-1-(7-3)(高等学校(通信制))", "高等学校", "通信制"),
    ("5-1-8(特別支援学校)", "特別支援学校", ""),
]

PUBLIC_SOURCE_NAME = "令和7年版千葉県教育便覧 V学校名簿 5-1 国・公立学校名簿"
PUBLIC_SOURCE_URL = "https://www.pref.chiba.lg.jp/kyouiku/seisaku/kouhou/kyouikubinran/documents/5-1-r7.xls"
PUBLIC_SOURCE_DATE = "2025-05-01"


def find_public_header(rows: list[list[Any]], max_scan: int = 10) -> tuple[int, dict[str, int]]:
    for row_index in range(min(max_scan, len(rows))):
        row = rows[row_index]
        column_map: dict[str, int] = {}
        for col_index, cell in enumerate(row):
            header = compact_header(cell)
            if not header:
                continue
            for field_name, aliases in PUBLIC_HEADER_ALIASES.items():
                if any(compact_header(alias) in header for alias in aliases):
                    column_map.setdefault(field_name, col_index)
        if "name" in column_map and "address" in column_map:
            return row_index, column_map
    raise ValueError("学校名・所在地の見出し行を検出できませんでした")


def cell_at(row: list[Any], column_map: dict[str, int], field_name: str) -> Any:
    index = column_map.get(field_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def read_public_workbook(path: Path) -> list[dict[str, Any]]:
    book = xlrd.open_workbook(str(path))
    records: list[dict[str, Any]] = []

    for sheet_name, school_type, course_label in PUBLIC_SHEETS:
        try:
            sheet = book.sheet_by_name(sheet_name)
        except xlrd.XLRDError:
            warn("public", f"シート未検出: {sheet_name}")
            continue

        rows = [sheet.row_values(r) for r in range(sheet.nrows)]
        if not rows:
            continue
        try:
            header_row, column_map = find_public_header(rows)
        except ValueError as exc:
            warn("public", f"{sheet_name}: {exc}")
            continue

        for row in rows[header_row + 1:]:
            name = normalize_name(cell_at(row, column_map, "name"))
            if not name:
                continue
            if any(word in name for word in ("計", "合計", "学校数")):
                continue

            committee = normalize_text(cell_at(row, column_map, "committee"))
            address = normalize_address(cell_at(row, column_map, "address"))
            if not address:
                continue
            establishment = "国立" if committee == "国立" else "公立"

            record = make_record(
                name=name,
                name_kana="",
                postal_code=cell_at(row, column_map, "postal"),
                address=address,
                school_type=school_type,
                establishment=establishment,
                operator=committee,
                phone=cell_at(row, column_map, "phone"),
                website="",
                source_name=PUBLIC_SOURCE_NAME,
                source_url=PUBLIC_SOURCE_URL,
                source_date=PUBLIC_SOURCE_DATE,
                course=[course_label] if course_label else [],
            )
            record["_course_label"] = course_label
            records.append(record)

    return records


def merge_public_high_schools(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """公立高等学校の全日制/定時制/通信制シートを1校1レコードへ統合する。"""
    high = [r for r in records if r["school_type"] == "高等学校"]
    others = [r for r in records if r["school_type"] != "高等学校"]

    groups: dict[str, list[dict[str, Any]]] = {}
    for record in high:
        key = f"{record['municipality']}|{record['name']}"
        groups.setdefault(key, []).append(record)

    merged: list[dict[str, Any]] = []
    for key, group in groups.items():
        if len(group) == 1:
            base = group[0]
            base.pop("_course_label", None)
            merged.append(base)
            continue

        addresses = {g["address"] for g in group}
        phones = {g["phone"] for g in group if g["phone"]}
        if len(addresses) > 1 or len(phones) > 1:
            warn(
                "public-high-merge",
                f"{group[0]['name']}: 課程間で所在地または電話番号が不一致のため自動統合せず個別レコードとして保持 "
                f"(addresses={addresses}, phones={phones})",
            )
            for g in group:
                g.pop("_course_label", None)
            merged.extend(group)
            continue

        base = dict(group[0])
        base["course"] = sorted(
            {g["_course_label"] for g in group if g["_course_label"]},
            key=lambda c: HIGH_SCHOOL_COURSE_ORDER.index(c) if c in HIGH_SCHOOL_COURSE_ORDER else 99,
        )
        base.pop("_course_label", None)
        stable_key = "|".join((base["establishment"], base["school_type"], base["municipality"], base["name"], ",".join(base["course"])))
        base["id"] = f"chiba-{slug(stable_key)}"
        merged.append(base)

    for record in others:
        record.pop("_course_label", None)

    return others + merged


# ---------------------------------------------------------------------------
# 補正資料（r8-gakkouichiran-2_henkou_ichiran.xls）
# ---------------------------------------------------------------------------

@dataclass
class CorrectionSet:
    additions: list[dict[str, str]] = field(default_factory=list)
    abolitions: list[dict[str, str]] = field(default_factory=list)
    address_changes: list[dict[str, str]] = field(default_factory=list)
    renames: list[dict[str, str]] = field(default_factory=list)
    suspensions: list[dict[str, str]] = field(default_factory=list)


def parse_correction_file(path: Path) -> CorrectionSet:
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]

    result = CorrectionSet()
    category = ""  # "elementary_junior" or "kindergarten"
    change_type = ""  # add / abolish / address_change / rename / suspend

    for row in rows:
        col0 = normalize_text(row[0]) if len(row) > 0 else ""
        col1 = normalize_text(row[1]) if len(row) > 1 else ""
        col2 = row[2] if len(row) > 2 else ""
        col3 = normalize_text(row[3]) if len(row) > 3 else ""

        is_marker_row = bool(col0) and not col1

        if is_marker_row:
            if "小学校" in col0 and "中学校" in col0:
                category = "elementary_junior"
            elif "幼稚園" in col0:
                category = "kindergarten"
            if "設置" in col0:
                change_type = "add"
            elif "廃止" in col0:
                change_type = "abolish"
            elif "所在地の変更" in col0:
                change_type = "address_change"
            elif "名称の変更" in col0:
                change_type = "rename"
            elif "休園" in col0:
                change_type = "suspend"
            continue

        if col0 in ("学校名", "園名"):
            continue

        if not col0 or not col1:
            continue

        name = normalize_name(col0)
        address = normalize_address(col1)
        date_iso = excel_serial_to_iso(col2)

        if change_type == "add":
            result.additions.append({
                "name": name, "address": address, "date": date_iso,
                "note": col3, "category": category,
            })
        elif change_type == "abolish":
            result.abolitions.append({
                "name": name, "address": address, "date": date_iso,
                "note": col3, "category": category,
            })
        elif change_type == "address_change":
            result.address_changes.append({
                "name": name, "new_address": address, "date": date_iso,
                "note": col3, "category": category,
            })
        elif change_type == "rename":
            # 備考: "○○小学校から変更" 形式から旧名称を抽出する
            old_name_match = re.search(r"(.+?)から変更", col3)
            old_name = normalize_name(old_name_match.group(1)) if old_name_match else ""
            result.renames.append({
                "old_name": old_name, "new_name": name, "address": address,
                "date": date_iso, "note": col3, "category": category,
            })
        elif change_type == "suspend":
            result.suspensions.append({
                "name": name, "address": address, "date": date_iso,
                "note": col3, "category": category,
            })

    return result


def infer_school_type_from_name(name: str, category: str) -> str:
    if category == "kindergarten":
        return "幼稚園"
    if name.endswith("中学校"):
        return "中学校"
    return "小学校"


def apply_corrections(records: list[dict[str, Any]], correction: CorrectionSet) -> list[dict[str, Any]]:
    by_name: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        by_name.setdefault(record["name"], []).append(record)

    removed_names: set[str] = set()

    for entry in correction.renames:
        old_name = entry["old_name"]
        new_name = entry["new_name"]
        matches = by_name.get(old_name, [])
        if not matches:
            warn("correction-rename", f"名称変更元 `{old_name}` が原本データに見つかりません（{entry['note']}）")
            continue
        for record in matches:
            record["name"] = new_name
            record["address"] = entry["address"] or record["address"]
            record["municipality"] = infer_municipality(record["address"])
        by_name.setdefault(new_name, []).extend(matches)

    for entry in correction.address_changes:
        matches = by_name.get(entry["name"], [])
        if not matches:
            warn("correction-address", f"所在地変更対象 `{entry['name']}` が原本データに見つかりません")
            continue
        for record in matches:
            record["address"] = entry["new_address"]
            record["municipality"] = infer_municipality(record["address"])
            warn(
                "correction-address",
                f"{entry['name']}: 所在地を `{entry['new_address']}` へ更新（郵便番号は原本のまま未検証）",
            )

    for entry in correction.abolitions:
        name = entry["name"]
        if name in by_name and by_name[name]:
            removed_names.add(name)
        else:
            warn("correction-abolish", f"廃止対象 `{name}` が原本データに見つかりません（{entry['note']}）")

    for entry in correction.suspensions:
        name = entry["name"]
        if name in by_name and by_name[name]:
            removed_names.add(name)
        else:
            warn("correction-suspend", f"休園対象 `{name}` が原本データに見つかりません")

    surviving = [r for r in records if r["name"] not in removed_names]

    for entry in correction.additions:
        name = entry["name"]
        if any(r["name"] == name for r in surviving):
            warn("correction-add", f"新設校 `{name}` は既に原本データに存在するため追加をスキップ")
            continue
        school_type = infer_school_type_from_name(name, entry["category"])
        record = make_record(
            name=name,
            name_kana="",
            postal_code="",
            address=entry["address"],
            school_type=school_type,
            establishment="公立",
            operator="",
            phone="",
            website="",
            source_name="公立学校一覧（令和7→8年度 変更一覧）",
            source_url="https://www.pref.chiba.lg.jp/kyouiku/seisaku/kouhou/gakkou-ichiran/documents/r8-gakkouichiran-2.xls",
            source_date="2026-04-24",
            course=[],
        )
        warn(
            "correction-add",
            f"新設校 `{name}`: 郵便番号・電話番号が変更一覧に記載がないため空欄で登録（追加の一次資料確認が必要）",
        )
        surviving.append(record)

    return surviving


# ---------------------------------------------------------------------------
# 私立学校名簿（.xlsx, openpyxl）
# ---------------------------------------------------------------------------

PRIVATE_HEADER_ALIASES = {
    "name": ("学校名",),
    "operator": ("学校設置者",),
    "postal": ("郵便番号",),
    "address": ("所在地",),
    "phone": ("電話番号",),
    "course": ("課程",),
}


def find_private_header(rows: list[tuple[Any, ...]], max_scan: int = 8) -> tuple[int, dict[str, int]]:
    for row_index in range(min(max_scan, len(rows))):
        row = rows[row_index]
        column_map: dict[str, int] = {}
        for col_index, cell in enumerate(row):
            header = compact_header(cell)
            if not header:
                continue
            for field_name, aliases in PRIVATE_HEADER_ALIASES.items():
                if any(compact_header(alias) in header for alias in aliases):
                    column_map.setdefault(field_name, col_index)
        if "name" in column_map and "address" in column_map:
            return row_index, column_map
    raise ValueError("学校名・所在地の見出し行を検出できませんでした")


def read_private_workbook(
    path: Path, *, school_type: str, source_name: str, source_url: str, source_date: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header_row, column_map = find_private_header(rows)

    records: list[dict[str, Any]] = []
    for row in rows[header_row + 1:]:
        name = normalize_name(cell_at(list(row), column_map, "name"))
        if not name:
            continue
        address = normalize_address(cell_at(list(row), column_map, "address"))
        if not address:
            continue

        course_raw = normalize_text(cell_at(list(row), column_map, "course"))
        course = [c for c in re.split(r"[\s　]+", course_raw) if c] if course_raw else []

        record = make_record(
            name=name,
            name_kana="",
            postal_code=cell_at(list(row), column_map, "postal"),
            address=address,
            school_type=school_type,
            establishment="私立",
            operator=cell_at(list(row), column_map, "operator"),
            phone=cell_at(list(row), column_map, "phone"),
            website="",
            source_name=source_name,
            source_url=source_url,
            source_date=source_date,
            course=course,
        )
        records.append(record)
    return records


SECONDARY_SUFFIX_RE = re.compile(r"(.*中等教育学校)\s*[（(](前期課程|後期課程)[）)]\s*$")


def split_secondary_school(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """名前が「○○中等教育学校（前期課程/後期課程）」の行を分離する。"""
    secondary: list[dict[str, Any]] = []
    plain: list[dict[str, Any]] = []
    for record in records:
        match = SECONDARY_SUFFIX_RE.match(record["name"])
        if match:
            record["_base_name"] = match.group(1)
            record["_stage"] = match.group(2)
            secondary.append(record)
        else:
            plain.append(record)
    return secondary, plain


def merge_private_secondary_schools(secondary: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for record in secondary:
        groups.setdefault(record["_base_name"], []).append(record)

    merged: list[dict[str, Any]] = []
    for base_name, group in groups.items():
        stages = {g["_stage"] for g in group}
        if len(group) != 2 or stages != {"前期課程", "後期課程"}:
            warn(
                "private-secondary-merge",
                f"{base_name}: 前期課程・後期課程が揃っていないため統合せず個別レコードとして保持 "
                f"(stages={stages})",
            )
            for g in group:
                g["school_type"] = "中等教育学校"
                g.pop("_base_name", None)
                g.pop("_stage", None)
            merged.extend(group)
            continue

        operators = {g["operator"] for g in group}
        addresses = {g["address"] for g in group}
        if len(operators) > 1 or len(addresses) > 1:
            warn(
                "private-secondary-merge",
                f"{base_name}: 前期課程・後期課程で設置者または所在地が不一致のため統合せず警告のみ "
                f"(operators={operators}, addresses={addresses})",
            )
            for g in group:
                g["school_type"] = "中等教育学校"
                g.pop("_base_name", None)
                g.pop("_stage", None)
            merged.extend(group)
            continue

        base = dict(group[0])
        base["name"] = base_name
        base["school_type"] = "中等教育学校"
        base["course"] = ["前期課程", "後期課程"]
        base.pop("_base_name", None)
        base.pop("_stage", None)
        stable_key = "|".join((base["establishment"], base["school_type"], base["municipality"], base["name"], ",".join(base["course"])))
        base["id"] = f"chiba-{slug(stable_key)}"
        merged.append(base)

    return merged


# ---------------------------------------------------------------------------
# 補正データ（公式一次資料で確認済みの新設校の郵便番号・電話番号等を補完する）
# ---------------------------------------------------------------------------

def apply_overrides(records: list[dict[str, Any]], overrides_path: Path) -> None:
    if not overrides_path.exists():
        return
    payload = json.loads(overrides_path.read_text(encoding="utf-8"))
    overrides = payload.get("overrides", {})
    by_name: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        by_name.setdefault(record["name"], []).append(record)

    for name, fields in overrides.items():
        matches = by_name.get(name, [])
        if not matches:
            warn("override", f"補正データ `{name}` に一致するレコードが見つかりません。")
            continue
        for record in matches:
            for key in ("postal_code", "address", "phone", "website", "source_url"):
                if fields.get(key):
                    record[key] = fields[key]
            record["municipality"] = infer_municipality(record["address"])
            warn("override", f"{name}: 公式一次資料により補完 ({fields.get('reason', '')})")


# ---------------------------------------------------------------------------
# メイン変換処理
# ---------------------------------------------------------------------------

def deduplicate(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id: dict[str, dict[str, Any]] = {}
    for record in records:
        record_id = record["id"]
        if record_id not in by_id:
            by_id[record_id] = record
            continue
        suffix = 2
        while f"{record_id}-{suffix}" in by_id:
            suffix += 1
        record["id"] = f"{record_id}-{suffix}"
        by_id[record["id"]] = record
    return list(by_id.values())


def sort_key(record: dict[str, Any]):
    m_idx = MUNICIPALITY_ORDER.index(record["municipality"]) if record["municipality"] in MUNICIPALITY_ORDER else 999
    t_idx = SCHOOL_TYPE_ORDER.index(record["school_type"]) if record["school_type"] in SCHOOL_TYPE_ORDER else 999
    return (m_idx, t_idx, record["name"])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, default=Path("data-source/chiba/2025"))
    parser.add_argument("--output", type=Path, default=Path("data/school-database/chiba.json"))
    parser.add_argument("--warnings-output", type=Path, default=Path("tools/school-database/chiba_conversion_warnings.json"))
    parser.add_argument("--overrides", type=Path, default=Path("tools/school-database/chiba_overrides.json"))
    args = parser.parse_args()

    # 1. 公立学校（単一ブック・複数シート）
    public_path = args.source_root / "5-1-r7_koritsu_meibo.xls"
    public_records = read_public_workbook(public_path)
    print(f"{public_path.name}: {len(public_records)} rows (merge前)")
    public_records = merge_public_high_schools(public_records)

    # 2. 補正（新設・廃止・所在地変更・名称変更・休園）
    correction_path = args.source_root / "r8-gakkouichiran-2_henkou_ichiran.xls"
    correction = parse_correction_file(correction_path)
    public_records = apply_corrections(public_records, correction)
    print(
        f"{correction_path.name}: 設置{len(correction.additions)} 廃止{len(correction.abolitions)} "
        f"所在地変更{len(correction.address_changes)} 名称変更{len(correction.renames)} 休園{len(correction.suspensions)}"
    )

    # 3. 私立学校（4ファイル）
    private_specs = [
        ("r8youchien_shiritsu.xlsx", "幼稚園",
         "私立幼稚園名簿（令和8年5月1日現在）",
         "https://www.pref.chiba.lg.jp/gakuji/shiritsutou/shiritsugakkou/documents/r8youchien.xlsx",
         "2026-05-01"),
        ("r8syougaku_shiritsu.xlsx", "小学校",
         "私立小学校名簿（令和8年5月1日現在）",
         "https://www.pref.chiba.lg.jp/gakuji/shiritsutou/shiritsugakkou/documents/r8syougaku.xlsx",
         "2026-05-01"),
        ("r8chugaku_shiritsu.xlsx", "中学校",
         "私立中学校・中等教育学校（前期）名簿（令和8年5月1日現在）",
         "https://www.pref.chiba.lg.jp/gakuji/shiritsutou/shiritsugakkou/documents/r8chugaku.xlsx",
         "2026-05-01"),
        ("r8koukou_shiritsu.xlsx", "高等学校",
         "私立高等学校・中等教育学校（後期）名簿（令和8年5月1日現在）",
         "https://www.pref.chiba.lg.jp/gakuji/shiritsutou/shiritsugakkou/documents/r8koukou.xlsx",
         "2026-05-01"),
    ]

    private_records: list[dict[str, Any]] = []
    secondary_candidates: list[dict[str, Any]] = []
    for filename, school_type, source_name, source_url, source_date in private_specs:
        path = args.source_root / filename
        if not path.exists():
            warn("private", f"SKIP missing: {path}")
            continue
        rows = read_private_workbook(
            path, school_type=school_type, source_name=source_name,
            source_url=source_url, source_date=source_date,
        )
        secondary, plain = split_secondary_school(rows)
        print(f"{filename}: {len(rows)} rows ({len(secondary)} 中等教育学校候補)")
        private_records.extend(plain)
        secondary_candidates.extend(secondary)

    private_records.extend(merge_private_secondary_schools(secondary_candidates))

    all_records = public_records + private_records
    apply_overrides(all_records, args.overrides)
    all_records = deduplicate(all_records)
    all_records.sort(key=sort_key)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(all_records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {len(all_records)} records to {args.output}")

    args.warnings_output.parent.mkdir(parents=True, exist_ok=True)
    warnings_payload = [{"context": w.context, "message": w.message} for w in WARNINGS]
    args.warnings_output.write_text(json.dumps(warnings_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(warnings_payload)} warnings to {args.warnings_output}")
    for w in WARNINGS:
        print(f"WARN [{w.context}] {w.message}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
