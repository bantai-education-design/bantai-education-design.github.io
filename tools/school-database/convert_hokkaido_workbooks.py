#!/usr/bin/env python3
"""北海道の公式HTML/PDF/Excel名簿を学校検索用JSONへ変換する。

原本は data-source/hokkaido/ に置く（Git管理外）。

北海道は14教育局が管内公立小中学校一覧をそれぞれ独自形式で公開しており
（HTML表・PDF罫線表・市町村別ページ等、局ごとに形式が異なる）、札幌市は
別途市教育委員会のページ、高等学校・特別支援学校は道教委が全道分をまとめた
PDF、私立学校は道庁が振興局×校種別のExcelを個別配布している。

このスクリプトは校種・情報源ごとに独立したパーサ関数を持ち、最後に
共通スキーマへ正規化してから統合・重複排除・JSON出力する。
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import pdfplumber
from bs4 import BeautifulSoup
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 正規化ユーティリティ
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\r", "")
    return text.strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("\n", "")
    return re.sub(r"[ \t　]+", "", text)


def normalize_postal_code(value: Any) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    return ""


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    text = normalize_text(value)
    text = text.replace("\n", "")
    text = text.replace("−", "-").replace("―", "-").replace("ー", "-").replace("－", "-")
    text = re.sub(r"[（(]代[）)]$", "", text)
    text = text.strip()
    if text in ("", "-", "なし", "同上"):
        return ""
    if re.fullmatch(r"\d{2,5}-\d{1,4}-\d{3,4}", text):
        return text
    return text


def combine_area_code(area_code: str, suffix: str) -> str:
    """ページ見出しの市外局番(例: 0135)と行側の残り桁(例: 22-3686)を結合する。"""
    area = re.sub(r"\D", "", normalize_text(area_code))
    suf = normalize_text(suffix).replace("−", "-").replace("―", "-")
    if not area or not suf:
        return normalize_phone(suffix)
    if re.fullmatch(r"\d{2,4}-\d{3,4}", suf):
        return f"{area}-{suf}"
    return normalize_phone(suffix)


def split_postal_and_address(value: Any) -> tuple[str, str]:
    text = normalize_text(value).replace("〒", "").replace("\n", "")
    match = re.match(r"^(\d{3})-?(\d{4})\s*(.*)$", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}", match.group(3).strip()
    return "", text


def normalize_address(value: Any, *, prefix: str = "") -> str:
    text = normalize_text(value)
    text = text.replace("\n", "")
    text = text.replace("−", "-").replace("―", "-").replace("－", "-")
    text = text.replace("〒", "")
    text = re.sub(r"^\d{3}-?\d{4}", "", text).strip()
    if prefix and text and prefix not in text:
        for i in range(len(prefix)):
            if text.startswith(prefix[i:]):
                text = prefix[:i] + text
                break
        else:
            text = prefix + text
    if text and not text.startswith("北海道"):
        text = "北海道" + text
    return text


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).lower()
    normalized = re.sub(r"[^0-9a-z぀-ヿ一-鿿]+", "-", normalized)
    return normalized.strip("-")


SOURCE_DATE = "2026-05-01"
PREF = "北海道"

records: list[dict[str, Any]] = []


def add_record(
    *,
    name: str,
    school_type: str,
    establishment: str,
    postal_code: str,
    address: str,
    municipality: str,
    phone: str,
    operator: str = "",
    website: str = "",
    course: list[str] | None = None,
    source_name: str,
    source_url: str,
) -> None:
    name = normalize_name(name)
    if not name:
        return
    rec = {
        "id": "",
        "prefecture": PREF,
        "name": name,
        "name_kana": "",
        "postal_code": normalize_postal_code(postal_code) if postal_code else "",
        "address": address,
        "municipality": municipality,
        "school_type": school_type,
        "establishment": establishment,
        "operator": operator,
        "phone": normalize_phone(phone),
        "website": website,
        "source_name": source_name,
        "source_url": source_url,
        "source_date": SOURCE_DATE,
        "verified_date": "",
        "course": course or [],
    }
    records.append(rec)


def infer_school_type(name: str) -> str:
    n = name
    if "義務教育学校" in n:
        return "義務教育学校"
    if "中等教育学校" in n:
        return "中等教育学校"
    if "認定こども園" in n or "こども園" in n:
        return "認定こども園"
    if "幼稚園" in n:
        return "幼稚園"
    if "特別支援" in n or "養護" in n or "盲学校" in n or "聾学校" in n:
        return "特別支援学校"
    if n.endswith("高") or "高等学校" in n or n.endswith("高校"):
        return "高等学校"
    if "中学校" in n or n.endswith("中"):
        return "中学校"
    if "小学校" in n or n.endswith("小"):
        return "小学校"
    return "小学校"


def cell_lines(cell) -> list[str]:
    """<br>区切りのtd/thセル内容を行のリストとして取得する。"""
    text = cell.get_text(separator="\n", strip=True)
    return [normalize_text(line) for line in text.split("\n") if normalize_text(line)]


DATA_DIR = Path(__file__).resolve().parents[2] / "data-source" / "hokkaido"


# ---------------------------------------------------------------------------
# 空知教育局: HTML表(rowspan方式、市町村名+学校名+郵便番号+住所+電話)
# ---------------------------------------------------------------------------

def parse_sorachi() -> None:
    path = DATA_DIR / "sorachi_shochu.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    table = soup.find_all("table")[0]
    rows = table.find_all("tr")[1:]
    muni = ""
    for row in rows:
        cells = row.find_all(["td", "th"])
        texts = [c.get_text(strip=True) for c in cells]
        if len(texts) == 5:
            muni_raw = texts[0]
            muni = re.sub(r"(小学校|中学校|義務教育学校)\d*", "", muni_raw).strip()
            name, postal, addr, phone = texts[1:]
        elif len(texts) == 4:
            name, postal, addr, phone = texts
        else:
            continue
        add_record(
            name=name,
            school_type=infer_school_type(name),
            establishment="公立",
            postal_code=postal,
            address=normalize_address(addr, prefix=muni),
            municipality=muni,
            phone=phone,
            source_name="空知教育局 公立小・中学校一覧",
            source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/stk/soragatukouitiran.html",
        )


# ---------------------------------------------------------------------------
# 日高教育局: 校種別テーブル、学校名に市町村名を含む(例: 日高町立日高小学校)
# ---------------------------------------------------------------------------

def parse_hidaka() -> None:
    path = DATA_DIR / "hidaka_gakkou.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    tables = soup.find_all("table")
    # table0=小学校 table1=中学校 table2=高校(skip, dedicated PDF優先) table3=特別支援(skip) table4=認定こども園
    for idx, table in enumerate(tables):
        if idx in (2, 3):
            continue
        rows = table.find_all("tr")[1:]
        for row in rows:
            cells = row.find_all(["td", "th"])
            texts = [c.get_text(strip=True) for c in cells]
            if len(texts) < 4:
                continue
            name, postal, addr, phone = texts[0], texts[1], texts[2], texts[3]
            m = re.match(r"^(.+?[市町村])立?", name)
            muni = m.group(1) if m else ""
            add_record(
                name=name,
                school_type="認定こども園" if idx == 4 else infer_school_type(name),
                establishment="公立",
                postal_code=postal,
                address=normalize_address(addr, prefix=muni),
                municipality=muni,
                phone=phone,
                source_name="日高教育局 学校一覧",
                source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/hdk/gakkouitiran.html",
            )


# ---------------------------------------------------------------------------
# 檜山教育局: 校種別テーブルに町名列あり(幼稚園/小学校/中学校)
# ---------------------------------------------------------------------------

def parse_hiyama() -> None:
    path = DATA_DIR / "hiyama_gakkou.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    tables = soup.find_all("table")
    # table2=幼稚園 table3=小学校 table4=中学校 (table0/1=統計, table5/6=道立高校・支援学校favor dedicated PDF)
    type_by_idx = {2: "幼稚園", 3: "小学校", 4: "中学校"}
    for idx, stype in type_by_idx.items():
        table = tables[idx]
        rows = table.find_all("tr")[1:]
        for row in rows:
            cells = row.find_all(["td", "th"])
            texts = [c.get_text(strip=True) for c in cells]
            if len(texts) < 5:
                continue
            muni, name, postal, addr, phone = texts[:5]
            add_record(
                name=name,
                school_type=stype,
                establishment="公立",
                postal_code=postal,
                address=normalize_address(addr, prefix=muni),
                municipality=muni,
                phone=phone,
                source_name="檜山教育局 学校一覧",
                source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/hyk/gakkouichiran.html",
            )


# ---------------------------------------------------------------------------
# 根室教育局: 一部テーブルは市町名列あり、一部は無し(直前の市町名を引き継ぐ)
# ---------------------------------------------------------------------------

def parse_nemuro() -> None:
    path = DATA_DIR / "nemuro_gakkou.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    elems = soup.find_all(["h2", "h3", "h4", "table"])
    current_muni = ""
    for el in elems:
        if el.name != "table":
            text = normalize_text(el.get_text())
            m = re.search(r"([^\s]+[市町村])", text)
            if m:
                current_muni = m.group(1)
            continue
        rows = el.find_all("tr")
        if not rows:
            continue
        header = [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])]
        has_muni_col = header and header[0] in ("市町名", "市町村名")
        has_kubun_col = header and header[0] == "設置区分"
        if has_kubun_col:
            continue  # 高校一覧は専用PDFを優先
        for row in rows[1:]:
            cells = row.find_all(["td", "th"])
            texts = [c.get_text(strip=True) for c in cells]
            if has_muni_col:
                if len(texts) < 5:
                    continue
                muni, name, postal, addr, phone = texts[:5]
                current_muni = muni
            else:
                if len(texts) < 4:
                    continue
                name, postal, addr, phone = texts[:4]
                muni = current_muni
            add_record(
                name=name,
                school_type=infer_school_type(name),
                establishment="公立",
                postal_code=postal,
                address=normalize_address(addr, prefix=muni),
                municipality=muni,
                phone=phone,
                source_name="根室教育局 管内公立学校一覧",
                source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/nky/kannnaikouritugaltukou.html",
            )


# ---------------------------------------------------------------------------
# 釧路教育局: 郵便番号・所在地と電話・FAXがそれぞれ1セルに結合。表順=市町村順。
# ---------------------------------------------------------------------------

KUSHIRO_MUNI_ORDER = ["釧路市", "釧路町", "厚岸町", "浜中町", "標茶町", "弟子屈町", "鶴居村", "白糠町"]


def parse_kushiro() -> None:
    path = DATA_DIR / "kushiro_gakkou.html"
    soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
    tables = soup.find_all("table")
    # table3-10 = 小中学校(市町村ごとに1表, KUSHIRO_MUNI_ORDER順) table11=高校(skip) table12=特別支援(skip) table13=認定こども園
    for i, muni in enumerate(KUSHIRO_MUNI_ORDER):
        table = tables[3 + i]
        rows = table.find_all("tr")[1:]
        for row in rows:
            cells = row.find_all(["td", "th"])
            texts = [c.get_text(strip=True) for c in cells]
            if len(texts) < 3:
                continue
            name, tel_fax, postal_addr = texts[:3]
            phone = re.split(r"ＦＡＸ|FAX", tel_fax)[0].strip()
            postal, addr = split_postal_and_address(postal_addr)
            add_record(
                name=name,
                school_type=infer_school_type(name),
                establishment="公立",
                postal_code=postal,
                address=normalize_address(addr, prefix=muni),
                municipality=muni,
                phone=phone,
                source_name="釧路教育局 学校一覧",
                source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/krk/page03/gakkouitiran.html",
            )
    # table13 = 認定こども園(釧路市)
    table = tables[13]
    rows = table.find_all("tr")[1:]
    for row in rows:
        cells = row.find_all(["td", "th"])
        texts = [c.get_text(strip=True) for c in cells]
        if len(texts) < 3:
            continue
        name, tel_fax, postal_addr = texts[:3]
        phone = re.split(r"ＦＡＸ|FAX", tel_fax)[0].strip()
        postal, addr = split_postal_and_address(postal_addr)
        add_record(
            name=name,
            school_type="認定こども園",
            establishment="公立",
            postal_code=postal,
            address=normalize_address(addr, prefix="釧路市"),
            municipality="釧路市",
            phone=phone,
            source_name="釧路教育局 学校一覧",
            source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/krk/page03/gakkouitiran.html",
        )


# ---------------------------------------------------------------------------
# 札幌市立小学校・中学校: 区ごとのテーブル(h3見出しで区名)。郵便番号なし。
# ---------------------------------------------------------------------------

def parse_sapporo() -> None:
    for fname, stype in [("sapporo_shogaku.html", "小学校"), ("sapporo_chugaku.html", "中学校")]:
        path = DATA_DIR / fname
        soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
        elems = soup.find_all(["h3", "table"])
        current_ward = ""
        for el in elems:
            if el.name == "h3":
                current_ward = normalize_text(el.get_text())
                continue
            rows = el.find_all("tr")[1:]
            for row in rows:
                cells = row.find_all(["td", "th"])
                texts = [c.get_text(strip=True) for c in cells]
                if len(texts) < 3:
                    continue
                short_name, addr, phone = texts[0], texts[1], texts[2]
                muni = f"札幌市{current_ward}"
                name = short_name if short_name.endswith(("小学校", "中学校", "分校")) else short_name + stype
                add_record(
                    name=name,
                    school_type=stype,
                    establishment="公立",
                    postal_code="",
                    address=normalize_address(addr, prefix=muni),
                    municipality=muni,
                    phone=phone,
                    source_name=f"札幌市立学校一覧-{stype}",
                    source_url=f"https://www.city.sapporo.jp/kyoiku/top/school/ichiran/{'shogaku' if stype=='小学校' else 'chugaku'}.html",
                )


# ---------------------------------------------------------------------------
# 胆振教育局: 市町村別11ページ、各列が<br>区切りで1セルに連結
#   住所列/電話列の先頭行はそれぞれ市町名/市外局番の共通プレフィックス
# ---------------------------------------------------------------------------

IBURI_MUNICIPALITIES = {
    "toyoura": "豊浦町", "toyako": "洞爺湖町", "sobetu": "壮瞥町", "shiraoi": "白老町",
    "abira": "安平町", "atuma": "厚真町", "mukawa": "むかわ町", "muroran": "室蘭市",
    "tomakomai": "苫小牧市", "noboribetu": "登別市", "date": "伊達市",
}


def parse_iburi() -> None:
    for fname, muni in IBURI_MUNICIPALITIES.items():
        path = DATA_DIR / "iburi" / f"{fname}.html"
        if not path.exists():
            continue
        soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
        tables = soup.find_all("table")
        for table in tables:
            rows = table.find_all("tr")
            if len(rows) < 2:
                continue
            header_cells = rows[0].find_all(["td", "th"])
            header = [normalize_text(c.get_text()) for c in header_cells]
            if not header or "学" not in header[0]:
                continue
            data_row = rows[1].find_all(["td", "th"])
            if len(data_row) < 4:
                continue
            names = cell_lines(data_row[0])
            addrs = cell_lines(data_row[1])
            postals = cell_lines(data_row[2])
            phones = cell_lines(data_row[3])
            if len(addrs) == len(names) + 1:
                addr_prefix = addrs[0]
                addrs = addrs[1:]
            else:
                addr_prefix = muni
            if len(phones) == len(names) + 1:
                area_code = phones[0]
                phones = phones[1:]
            else:
                area_code = ""
            for i, name in enumerate(names):
                addr = addrs[i] if i < len(addrs) else ""
                postal = postals[i] if i < len(postals) else ""
                phone_raw = phones[i] if i < len(phones) else ""
                phone = combine_area_code(area_code, phone_raw) if area_code else normalize_phone(phone_raw)
                full_addr = addr if addr.startswith(muni) else muni + addr
                add_record(
                    name=name,
                    school_type=infer_school_type(name),
                    establishment="公立",
                    postal_code=postal,
                    address=normalize_address(full_addr),
                    municipality=muni,
                    phone=phone,
                    source_name=f"胆振教育局 学校一覧({muni})",
                    source_url=f"https://www.dokyoi.pref.hokkaido.lg.jp/hk/ibk/ichiran_{fname}.html",
                )


# ---------------------------------------------------------------------------
# 石狩教育局・高校・中等教育学校: pdfplumberのextract_textがそのまま整った
# 1行1校のテキストになるPDF(iTextSharp出力)。正規表現で行を分解する。
# ---------------------------------------------------------------------------

def parse_ishikari() -> None:
    path = DATA_DIR / "ishikari_koritsu_gakkou_ichiran.pdf"
    muni = ""
    stype = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or len(row) < 13:
                        continue
                    muni_cell, type_cell, name_cell = row[0], row[1], row[2]
                    addr_cell, postal_cell, phone_cell = row[10], row[11], row[12]
                    if muni_cell:
                        muni = normalize_name(muni_cell)
                    if type_cell:
                        stype = normalize_name(type_cell)
                    name = normalize_name(name_cell or "")
                    if not name or name in ("計",) or not postal_cell or not re.match(r"\d{3}-\d{4}", postal_cell):
                        continue
                    add_record(
                        name=name,
                        school_type=stype or infer_school_type(name),
                        establishment="公立",
                        postal_code=postal_cell,
                        address=normalize_address(addr_cell or "", prefix=muni),
                        municipality=muni,
                        phone=phone_cell or "",
                        source_name="石狩教育局 公立学校等一覧",
                        source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/ikk/ishidata.html",
                    )


def parse_koukou_and_chuto() -> None:
    path = DATA_DIR / "koritsu_koukou_ichiran.pdf"
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            section = "高等学校" if "高 等 学 校" in text or "高等学校" in text.split("\n")[0] else None
            for line in text.split("\n"):
                line = normalize_text(line)
                if "中 等 教 育 学 校" in line or "中等教育学校" in line:
                    section = "中等教育学校"
                    continue
                m = re.match(
                    r"^\d+\s+(?P<name>\S+?)(\s*※)?\s+(?P<postal>\d{3}-\d{4})\s+(?P<addr>\S+)\s+(?P<phone>[\d-]+)$",
                    line,
                )
                if not m:
                    continue
                establishment = "市町村立" if "※" in line else "道立"
                add_record(
                    name=m.group("name"),
                    school_type=section or "高等学校",
                    establishment=establishment,
                    postal_code=m.group("postal"),
                    address=normalize_address(m.group("addr")),
                    municipality="",
                    phone=m.group("phone"),
                    source_name="北海道公立高等学校・中等教育学校一覧",
                    source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/kki/gakkou.html",
                )


# ---------------------------------------------------------------------------
# 後志教育局: 罫線表(extract_tables)。市町村名は先頭列にrowspan、電話は
# ページ見出し「電話番号（０１３５）」の市外局番+行側4桁で構成される。
# ---------------------------------------------------------------------------

AREA_CODE_HEADER_RE = re.compile(r"電話番号\s*[\(（](\d{2,4})[\)）]")


def clean_school_name(raw: str) -> str:
    return normalize_name(re.sub(r"\d+$", "", raw))


def parse_shiribeshi() -> None:
    path = DATA_DIR / "shiribeshi_shochu.pdf"
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            m = AREA_CODE_HEADER_RE.search(text)
            area_code = m.group(1) if m else ""
            tables = page.extract_tables()
            if not tables:
                continue
            table = tables[0]
            muni = ""
            for row in table[1:]:
                if not row or len(row) < 11:
                    continue
                muni_cell, name_cell = row[0], row[1]
                if muni_cell:
                    muni = normalize_name(muni_cell)
                name = clean_school_name(name_cell or "")
                if not name or name in ("計",):
                    continue
                addr = row[8] or ""
                phone_suffix = row[9] or ""
                phone = combine_area_code(area_code, phone_suffix) if area_code else normalize_phone(phone_suffix)
                add_record(
                    name=name,
                    school_type=infer_school_type(name),
                    establishment="公立",
                    postal_code="",
                    address=normalize_address(addr, prefix=muni),
                    municipality=muni,
                    phone=phone,
                    source_name="後志教育局 公立小学校・中学校一覧",
                    source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/sbk/youran.html",
                )


# ---------------------------------------------------------------------------
# 宗谷教育局: 罫線表、3行1校(1行目=本体、2/3行目=URL等の付随行)。
# 「電話番号等」列に〒郵便番号+住所+改行+電話が1セルで入っている。
# ---------------------------------------------------------------------------

def parse_soya() -> None:
    for path, default_type in [
        (DATA_DIR / "soya_shochu.pdf", None),
        (DATA_DIR / "soya_koutokushi.pdf", None),
    ]:
        with pdfplumber.open(path) as pdf:
            muni = ""
            for page in pdf.pages:
                tables = page.extract_tables()
                if not tables:
                    continue
                for table in tables:
                    for row in table:
                        if not row or len(row) < 9:
                            continue
                        muni_cell, name_cell = row[0], row[1]
                        if muni_cell:
                            muni = normalize_name(muni_cell)
                        name = normalize_name(name_cell or "")
                        if not name or name in ("学校名", "") or "一覧" in name:
                            continue
                        combined = row[8] or ""
                        if "〒" not in combined:
                            continue
                        postal, rest = split_postal_and_address(combined)
                        parts = rest.split("\n") if "\n" in rest else [rest]
                        addr = parts[0]
                        phone = ""
                        pm = re.search(r"(\d{2,5}-\d{1,4}-\d{3,4})", combined)
                        if pm:
                            phone = pm.group(1)
                        add_record(
                            name=name,
                            school_type=infer_school_type(name),
                            establishment="公立",
                            postal_code=postal,
                            address=normalize_address(addr, prefix=muni),
                            municipality=muni,
                            phone=phone,
                            source_name="宗谷教育局 学校一覧",
                            source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/syk/156121.html",
                        )


# ---------------------------------------------------------------------------
# オホーツク教育局: 罫線表、住所列/電話列がそれぞれ〒+改行+住所 / 電話+改行+FAX
# ---------------------------------------------------------------------------

def parse_okhotsk() -> None:
    # 高校・特別支援学校を含む okhotsk_koutokushi.pdf は道教委の全道一覧
    # (koritsu_koukou_ichiran.pdf / 特別支援学校所在地等一覧)と重複するため対象外。
    for path in [DATA_DIR / "okhotsk_kodomoen.pdf", DATA_DIR / "okhotsk_shochu.pdf"]:
        with pdfplumber.open(path) as pdf:
            muni = ""
            establishment_hint = "公立"
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row or len(row) < 8:
                            continue
                        muni_cell, name_cell = row[0], row[1]
                        if muni_cell:
                            muni = normalize_name(muni_cell)
                        name = normalize_name(name_cell or "")
                        if not name or name in ("学校名", "園名") or "一覧" in name:
                            continue
                        addr_cell = row[-2] or ""
                        phone_cell = row[-1] or ""
                        postal, addr = split_postal_and_address(addr_cell)
                        phone = (phone_cell.split("\n")[0] if phone_cell else "").strip()
                        add_record(
                            name=name,
                            school_type=infer_school_type(name),
                            establishment=establishment_hint,
                            postal_code=postal,
                            address=normalize_address(addr, prefix=muni),
                            municipality=muni,
                            phone=phone,
                            source_name="オホーツク教育局 学校一覧",
                            source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/okh/school.html",
                        )


# ---------------------------------------------------------------------------
# 上川教育局・十勝教育局: 要覧PDF中の学校一覧ページ(罫線表)
# ---------------------------------------------------------------------------

def parse_kamikawa() -> None:
    path = DATA_DIR / "kamikawa_youran.pdf"
    with pdfplumber.open(path) as pdf:
        muni = ""
        for i in range(9, 14):
            if i >= len(pdf.pages):
                break
            page = pdf.pages[i]
            tables = page.extract_tables()
            for table in tables:
                for row in table[1:] if table else []:
                    if not row or len(row) < 10:
                        continue
                    muni_cell, type_cell, name_cell = row[0], row[1], row[2]
                    addr_cell, phone_cell = row[8], row[9]
                    if muni_cell:
                        muni = normalize_name(muni_cell)
                    name = normalize_name(name_cell or "")
                    if not name or "学校名" in name or name == "計":
                        continue
                    postal, addr = split_postal_and_address(addr_cell or "")
                    if not postal:
                        continue
                    phone = (phone_cell or "").split("\n")[0].strip() if phone_cell else ""
                    add_record(
                        name=name,
                        school_type=infer_school_type(name),
                        establishment="公立",
                        postal_code=postal,
                        address=normalize_address(addr, prefix=muni),
                        municipality=muni,
                        phone=normalize_phone(phone),
                        source_name="上川教育局要覧「上川の教育2026」学校一覧",
                        source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/kkk/kamikawayouran.html",
                    )


def parse_tokachi() -> None:
    path = DATA_DIR / "tokachi_youran.pdf"
    with pdfplumber.open(path) as pdf:
        muni = ""
        for i in range(16, 20):
            if i >= len(pdf.pages):
                break
            page = pdf.pages[i]
            tables = page.extract_tables()
            for table in tables:
                for row in table[1:] if table else []:
                    if not row or len(row) < 6:
                        continue
                    muni_cell, name_cell, addr_cell = row[0], row[1], row[2]
                    if muni_cell:
                        muni = normalize_name(muni_cell)
                    name = normalize_name(name_cell or "")
                    if not name or name in ("学校名",) or "高等学校" in name:
                        continue
                    postal, addr = split_postal_and_address(addr_cell or "")
                    phone_cell = row[3] if len(row) > 3 else ""
                    add_record(
                        name=name,
                        school_type=infer_school_type(name),
                        establishment="公立",
                        postal_code=postal,
                        address=normalize_address(addr, prefix=muni),
                        municipality=muni,
                        phone=normalize_phone(phone_cell),
                        source_name="十勝教育局要覧「十勝の教育」学校一覧",
                        source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/tky/65283.html",
                    )


# ---------------------------------------------------------------------------
# 渡島教育局要覧(2/3分冊)・留萌教育局要覧
# ---------------------------------------------------------------------------

def parse_oshima() -> None:
    path = DATA_DIR / "oshima_youran_2.pdf"
    with pdfplumber.open(path) as pdf:
        # page2(index)=幼稚園: [1]=市町名 [2]=学校名 [8]=住所 [9]=電話
        # page3-7(index)=小中学校: [1]=市町名 [2]=学校種別 [3]=学校名 [11]=住所 [12]=電話
        muni = ""
        stype = ""
        page = pdf.pages[2]
        for table in page.extract_tables():
            for row in table:
                if not row or len(row) < 10:
                    continue
                muni_cell, name_cell, addr_cell, phone_cell = row[1], row[2], row[8], row[9]
                if muni_cell:
                    muni = normalize_name(muni_cell)
                name = normalize_name(name_cell or "")
                if not name or not addr_cell or "〒" not in (addr_cell or ""):
                    continue
                postal, addr = split_postal_and_address(addr_cell)
                add_record(
                    name=name,
                    school_type="幼稚園",
                    establishment="公立",
                    postal_code=postal,
                    address=normalize_address(addr, prefix=muni),
                    municipality=muni,
                    phone=(phone_cell or "").split("\n")[0],
                    source_name="渡島教育局要覧「渡島の教育」学校一覧(幼稚園)",
                    source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/oky/11_soumu/125972.html",
                )
        for i in range(3, 8):
            if i >= len(pdf.pages):
                break
            page = pdf.pages[i]
            for table in page.extract_tables():
                for row in table:
                    if not row or len(row) < 13:
                        continue
                    muni_cell, type_cell, name_cell = row[1], row[2], row[3]
                    addr_cell, phone_cell = row[11], row[12]
                    if muni_cell:
                        muni = normalize_name(muni_cell)
                    if type_cell:
                        stype = normalize_name(type_cell)
                    name = normalize_name(name_cell or "")
                    if not name or not addr_cell or "〒" not in (addr_cell or ""):
                        continue
                    postal, addr = split_postal_and_address(addr_cell)
                    add_record(
                        name=name,
                        school_type=stype or infer_school_type(name),
                        establishment="公立",
                        postal_code=postal,
                        address=normalize_address(addr, prefix=muni),
                        municipality=muni,
                        phone=(phone_cell or "").split("\n")[0],
                        source_name="渡島教育局要覧「渡島の教育」学校一覧",
                        source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/oky/11_soumu/125972.html",
                    )


def parse_rumoi() -> None:
    path = DATA_DIR / "rumoi_youran_p1_12.pdf"
    with pdfplumber.open(path) as pdf:
        muni = ""
        for i in [10, 11]:
            if i >= len(pdf.pages):
                break
            page = pdf.pages[i]
            tables = page.extract_tables()
            for table in tables:
                for row in table[1:] if table else []:
                    if not row:
                        continue
                    text_row = [normalize_text(c or "") for c in row]
                    joined = "".join(text_row)
                    if "〒" not in joined:
                        continue
                    name = ""
                    for cand in text_row:
                        cand_n = normalize_name(cand)
                        if cand_n and "〒" not in cand_n and not re.match(r"^\d", cand_n) and len(cand_n) <= 12 and cand_n not in ("併",):
                            name = re.sub(r"（併）$", "", cand_n)
                            break
                    if not name:
                        continue
                    postal_m = re.search(r"〒(\d{3}-?\d{4})", joined)
                    postal = normalize_postal_code(postal_m.group(1)) if postal_m else ""
                    addr_m = re.search(r"(留萌市|苫前郡\S+町|天塩郡\S+町|増毛郡増毛町|留萌郡小平町|羽幌町)\S+", joined)
                    addr = addr_m.group(0) if addr_m else ""
                    phone_m = re.search(r"\((\d{2,5})\)\s*([\d-]{7,9})", joined) or re.search(r"(\d{2,5}-\d{2,4}-\d{3,4})", joined)
                    if phone_m and phone_m.lastindex == 2:
                        phone = f"{phone_m.group(1)}-{phone_m.group(2)}"
                    elif phone_m:
                        phone = phone_m.group(1)
                    else:
                        phone = ""
                    add_record(
                        name=name,
                        school_type=infer_school_type(name),
                        establishment="公立",
                        postal_code=postal,
                        address=normalize_address(addr),
                        municipality="",
                        phone=phone,
                        source_name="留萌教育局要覧 小学校・中学校一覧",
                        source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/rky/youran.html",
                    )


# ---------------------------------------------------------------------------
# 特別支援学校: 所在地等一覧PDF(画像PDF)を目視で書き起こしたTSV
# ---------------------------------------------------------------------------

ESTABLISHMENT_MAP = {"道立": "道立", "国立": "国立", "市立": "市立", "私立": "私立"}


def parse_tokubetsu_shien() -> None:
    path = DATA_DIR / "tokubetsu_shien_transcribed.tsv"
    lines = path.read_text(encoding="utf-8").splitlines()
    header = lines[0].split("\t")
    for line in lines[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        row = dict(zip(header, cols))
        add_record(
            name=row.get("学校名", ""),
            school_type="特別支援学校",
            establishment=ESTABLISHMENT_MAP.get(row.get("設置者", ""), row.get("設置者", "")),
            postal_code=row.get("郵便番号", ""),
            address=normalize_address(row.get("所在地", "")),
            municipality=row.get("管内", ""),
            phone=row.get("電話番号", ""),
            source_name="北海道特別支援学校所在地等一覧",
            source_url="https://www.dokyoi.pref.hokkaido.lg.jp/hk/tkk/yoran.html",
        )


# ---------------------------------------------------------------------------
# 国立学校: 北海道教育大学附属学校園一覧(HTML、大学サイト)
# ---------------------------------------------------------------------------

NATIONAL_SCHOOLS = [
    ("北海道教育大学附属札幌小学校", "060-0011", "札幌市中央区北11条西16丁目", "札幌市中央区", "011-778-0471", "小学校"),
    ("北海道教育大学附属札幌中学校", "060-0011", "札幌市中央区北11条西16丁目", "札幌市中央区", "011-778-0481", "中学校"),
    ("北海道教育大学附属旭川幼稚園", "070-8006", "旭川市春光町6条9丁目", "旭川市", "0166-54-3556", "幼稚園"),
    ("北海道教育大学附属旭川小学校", "070-8006", "旭川市春光町6条9丁目", "旭川市", "0166-52-2361", "小学校"),
    ("北海道教育大学附属旭川中学校", "070-8006", "旭川市春光町6条9丁目", "旭川市", "0166-53-2751", "中学校"),
    ("北海道教育大学附属釧路義務教育学校", "085-0802", "釧路市桜ヶ岡3丁目1-1", "釧路市", "0154-91-6322", "義務教育学校"),
    ("北海道教育大学附属函館幼稚園", "040-8567", "函館市八幡町1-2", "函館市", "0138-46-2237", "幼稚園"),
    ("北海道教育大学附属函館小学校", "040-8567", "函館市八幡町1-2", "函館市", "0138-46-2235", "小学校"),
    ("北海道教育大学附属函館中学校", "040-8567", "函館市八幡町1-2", "函館市", "0138-46-2233", "中学校"),
]


def parse_national() -> None:
    for name, postal, addr, muni, phone, stype in NATIONAL_SCHOOLS:
        add_record(
            name=name,
            school_type=stype,
            establishment="国立",
            postal_code=postal,
            address=normalize_address(addr, prefix=muni),
            municipality=muni,
            phone=phone,
            source_name="北海道教育大学 附属学校園一覧",
            source_url="https://www.hokkyodai.ac.jp/attached/school_list/",
        )


# ---------------------------------------------------------------------------
# 私立学校: 振興局×校種別Excel(58ファイル)
# ---------------------------------------------------------------------------

PRIVATE_TYPE_MAP = {
    "yochien": "幼稚園",
    "ele": "小学校",
    "juni": "中学校",
    "high": "高等学校",
    "tsushin": "高等学校",
    "ktsushin": "高等学校",
    "tokubetsu": "特別支援学校",
    "sen": "専修学校",
    "kaku": "各種学校",
}


def parse_private_excel() -> None:
    private_dir = DATA_DIR / "private"
    if not private_dir.exists():
        return
    for path in sorted(private_dir.iterdir()):
        if path.suffix.lower() not in (".xlsx", ".xls"):
            continue
        prefix = path.stem.split("_")[0]
        stype = PRIVATE_TYPE_MAP.get(prefix, "")
        if not stype:
            continue
        try:
            wb = load_workbook(path, data_only=True)
        except Exception:
            continue
        for ws in wb.worksheets:
            current_muni = ""
            header_seen = False
            for row in ws.iter_rows(values_only=True):
                cells = [c for c in row]
                texts = [normalize_text(c) if c is not None else "" for c in cells]
                joined = "".join(texts)
                if not joined:
                    continue
                if texts[0].startswith("●"):
                    current_muni = texts[0].lstrip("●")
                    continue
                if "名" in texts[0] and ("所" in "".join(texts[1:3]) or "地" in "".join(texts[1:3])):
                    header_seen = True
                    continue
                if not header_seen:
                    continue
                name = texts[0]
                if not name or name in ("計",):
                    continue
                addr = texts[1] if len(texts) > 1 else ""
                phone = texts[2] if len(texts) > 2 else ""
                if not re.search(r"\d{2,5}-\d{2,4}-\d{3,4}", phone):
                    continue
                website = texts[4] if len(texts) > 4 and texts[4].startswith("http") else ""
                add_record(
                    name=name,
                    school_type=stype,
                    establishment="私立",
                    postal_code="",
                    address=normalize_address(addr),
                    municipality="",
                    phone=phone,
                    website=website,
                    source_name=f"北海道私立学校一覧({stype})",
                    source_url="https://www.pref.hokkaido.lg.jp/sm/gkj/allschoolseach.html",
                )


# ---------------------------------------------------------------------------
# メイン処理: 統合・重複除外・ID付与・出力
# ---------------------------------------------------------------------------

MUNI_FROM_ADDRESS_RE = re.compile(r"^北海道(?:[^0-9〒]*?郡)?(?P<muni>[^0-9〒]+?[市町村](?:[^0-9〒]+?区)?)")


BAD_MUNI_PREFIX_RE = re.compile(r"^(立|道|市町村|認定こども園|大字|以下)")


def fill_missing_municipality() -> None:
    """住所文字列から市町村名を再抽出し、パーサ側で取り違えた設置区分の
    混入等(例:「市町村立」「認定こども園○○町」)を上書き修正する。
    住所から抽出できない場合のみ、既存の値を保持する。"""
    for rec in records:
        m = MUNI_FROM_ADDRESS_RE.match(rec["address"])
        if m:
            rec["municipality"] = m.group("muni")
        if len(rec["municipality"]) <= 1 or BAD_MUNI_PREFIX_RE.match(rec["municipality"]):
            rec["municipality"] = ""


FULL_NAME_SUFFIXES = (
    "幼稚園", "小学校", "中学校", "義務教育学校", "高等学校", "中等教育学校",
    "特別支援学校", "認定こども園", "学園", "学院",
)

SINGLE_CHAR_SUFFIX = {"小学校": "小", "中学校": "中"}

NAME_BUILD_TYPES = {"幼稚園", "小学校", "中学校", "義務教育学校", "高等学校", "中等教育学校"}

COURSE_QUALIFIER_RE = re.compile(r"[（(][^（）()]*[）)]\s*$")


def strip_course_qualifier(name: str) -> str:
    """「岩見沢東（全日制）」のような課程・分校の注記を除去する(正式名称には含めない)。"""
    return COURSE_QUALIFIER_RE.sub("", name).strip()


def looks_like_official_name(name: str) -> bool:
    return "立" in name and any(name.endswith(s) for s in FULL_NAME_SUFFIXES)


def municipal_legal_name(municipality: str) -> str:
    """学校の正式名称の設置者表記では、札幌市は区名を含めず「札幌市」とする。"""
    if municipality.startswith("札幌市"):
        return "札幌市"
    return municipality


INSTITUTION_PREFIX_RE = re.compile(r"^(市立|町立|村立|道立|国立)")

NATIONAL_UNIV_MARKER = "北海道教育大学附属"


def fix_misclassified_establishment() -> None:
    """教育局のHTML表に他設置者の学校が紛れ込んでいた場合の設置区分を補正する。
    (例: 釧路教育局の一覧に混在する私立中学校、国立大学附属校の重複行)"""
    fixed: list[dict[str, Any]] = []
    for rec in records:
        name = rec["name"]
        if NATIONAL_UNIV_MARKER in name and rec["establishment"] != "国立":
            # 北海道教育大学附属校はparse_national()側で正式名称を別途収録済みのため、
            # 各教育局の一覧に紛れ込んだ重複行はここで除外する。
            continue
        if name.startswith("私立") and rec["establishment"] not in ("私立",):
            rec["establishment"] = "私立"
            rec["name"] = name[len("私立"):]
        fixed.append(rec)
    records.clear()
    records.extend(fixed)


def collapse_repeated_block(name: str) -> str:
    """「茶路中学校茶路中学校」のように原本側の抽出崩れで名称全体が連続して
    重複した場合に、後半の重複ブロックを取り除く。"""
    n = len(name)
    for block_len in range(2, n // 2 + 1):
        if name[-block_len:] == name[-2 * block_len : -block_len]:
            return name[:-block_len]
    return name


def build_official_names() -> None:
    """略称のまま収録された学校名を「{設置者}立{校名}{校種}」の正式名称に組み立てる。
    原本の表記がすでに正式名称の場合は変更しない。私立・国立は原本が正式名称の
    ため対象外。特別支援学校は別途手作業で正式名称を収録済みのため対象外。"""
    for rec in records:
        establishment = rec["establishment"]
        school_type = rec["school_type"]
        if establishment in ("私立", "国立"):
            continue
        if school_type not in NAME_BUILD_TYPES:
            continue

        name = strip_course_qualifier(rec["name"])

        if looks_like_official_name(name):
            rec["name"] = name
            continue

        # 原本側にすでに「市立」等の設置者表記のみが付いている場合(例:「市立札幌新川」)は
        # 一旦取り除き、establishmentに基づく正しい設置者表記で組み直す。
        name = INSTITUTION_PREFIX_RE.sub("", name)

        full_suffix = school_type
        suffix = full_suffix
        if name.endswith(full_suffix):
            core = name[: -len(full_suffix)]
        elif name.endswith("分校"):
            # 「○○中学校△△分校」のように分校名で完結する場合は校種を重ねて付けない。
            core = name
            suffix = ""
        else:
            single = SINGLE_CHAR_SUFFIX.get(school_type)
            if single and name.endswith(single):
                core = name[:-1]
            elif school_type == "義務教育学校" and name.endswith("校"):
                # 「○○学校」のように末尾が「校」で完結する名称も多いため、
                # その場合は「義務教育学校」を重ねて付けない。
                core = name
                suffix = ""
            else:
                core = name

        if not core:
            continue

        if school_type in ("高等学校", "中等教育学校"):
            prefix = "北海道" if establishment == "道立" else municipal_legal_name(rec["municipality"])
        else:
            prefix = municipal_legal_name(rec["municipality"])

        rec["name"] = f"{prefix}立{core}{suffix}" if prefix else f"{core}{suffix}"


ESTABLISHMENT_NORMALIZE = {
    "道立": "公立",
    "市町村立": "公立",
    "市立": "公立",
    "公立": "公立",
    "私立": "私立",
    "国立": "国立",
}

ALLOWED_SCHOOL_TYPES = {
    "幼稚園",
    "幼保連携型認定こども園",
    "小学校",
    "中学校",
    "義務教育学校",
    "高等学校",
    "中等教育学校",
    "特別支援学校",
}

DROP_SCHOOL_TYPES = {"専修学校", "各種学校"}

BAD_ADDRESS_TOKENS = ("市町村名", "学校名", "学校種別", "小学校函館", "中学校函館")

BARE_TYPE_NAMES = {
    "幼稚園", "小学校", "中学校", "義務教育学校", "高等学校", "中等教育学校", "特別支援学校",
}


def clean_records() -> None:
    cleaned: list[dict[str, Any]] = []
    for rec in records:
        if rec["school_type"] in DROP_SCHOOL_TYPES:
            continue
        if rec["name"] in BARE_TYPE_NAMES:
            # 校種名のみが学校名として残った抽出崩れの行は除外する
            continue
        if rec["school_type"] == "認定こども園":
            rec["school_type"] = "幼保連携型認定こども園"
        rec["establishment"] = ESTABLISHMENT_NORMALIZE.get(rec["establishment"], rec["establishment"])
        if rec["school_type"] not in ALLOWED_SCHOOL_TYPES:
            continue
        if rec["establishment"] not in ("公立", "私立", "国立"):
            continue
        if any(tok in rec["address"] for tok in BAD_ADDRESS_TOKENS):
            continue
        if not re.search(r"[市町村]", rec["address"]) and not rec["address"].startswith("北海道札幌市"):
            # 住所に市町村相当の文字列が無い(パース崩れ)ものは除外
            if len(rec["address"]) < 6:
                continue
        cleaned.append(rec)
    records.clear()
    records.extend(cleaned)


KANJI_DIGITS = "〇一二三四五六七八九"


def int_to_kanji(n: int) -> str:
    if n < 10:
        return KANJI_DIGITS[n]
    tens, ones = divmod(n, 10)
    s = ("" if tens == 1 else KANJI_DIGITS[tens]) + "十" if tens else ""
    return s + (KANJI_DIGITS[ones] if ones else "")


def arabicize_jo_to_kanji(text: str) -> str:
    """「北4条東」「あいの里2条」のような算用数字表記を、郵便番号データの
    漢数字表記(「北四条東」「あいの里二条」)に変換したバージョンを返す
    (住所欄自体は変更しない)。"""
    def repl(m: re.Match) -> str:
        return int_to_kanji(int(m.group(1))) + "条"

    return re.sub(r"(\d{1,2})条", repl, text)


def extract_chome(text: str) -> int | None:
    m = re.search(r"(\d{1,3})丁目", text)
    return int(m.group(1)) if m else None


def fill_missing_postal() -> None:
    """日本郵便の郵便番号データ(市区町村×町域)から、郵便番号欠落レコードを補完する。"""
    lookup_path = DATA_DIR / "ken_all_hokkaido.json"
    if not lookup_path.exists():
        return
    muni_town_postal: dict[str, list[list[str]]] = json.loads(lookup_path.read_text(encoding="utf-8"))
    for rec in records:
        if rec["postal_code"] or not rec["municipality"]:
            continue
        towns = muni_town_postal.get(rec["municipality"])
        if not towns:
            continue
        addr_tail = rec["address"]
        if addr_tail.startswith("北海道"):
            addr_tail = addr_tail[len("北海道"):]
        addr_tail = re.sub(r"^[^0-9〒]*?郡(?=[^0-9〒]+[市町村])", "", addr_tail)
        if addr_tail.startswith(rec["municipality"]):
            addr_tail = addr_tail[len(rec["municipality"]):]
        addr_tail = re.sub(r"^字", "", addr_tail)
        addr_tail_kanji = arabicize_jo_to_kanji(addr_tail)
        chome = extract_chome(addr_tail)
        best_match = None
        best_len = 0
        best_range: tuple[int, int] | None = None
        for town, postal in towns:
            range_m = re.search(r"[（(]\s*(\d+)\s*[〜~]\s*(\d+)\s*丁目\s*[）)]", town)
            town_clean = re.sub(r"[（(].*?[）)]", "", town)
            if town_clean in ("以下に掲載がない場合", ""):
                continue
            if not (addr_tail.startswith(town_clean) or addr_tail_kanji.startswith(town_clean)):
                continue
            if len(town_clean) < best_len:
                continue
            if range_m and chome is not None:
                lo, hi = int(range_m.group(1)), int(range_m.group(2))
                if not (lo <= chome <= hi):
                    continue
                best_match, best_len, best_range = postal, len(town_clean), (lo, hi)
            elif len(town_clean) > best_len or (len(town_clean) == best_len and best_range is None):
                best_match, best_len = postal, len(town_clean)
        if best_match:
            rec["postal_code"] = normalize_postal_code(best_match)
        elif len(towns) == 1:
            rec["postal_code"] = normalize_postal_code(towns[0][1])


def drop_incomplete_records() -> None:
    """住所からの再抽出後も市町村名が確定できなかったレコード(PDFのページ境界
    にまたがる行など、ごく少数の抽出崩れ)は必須項目未充足のため除外する。"""
    kept = [rec for rec in records if rec["municipality"]]
    records.clear()
    records.extend(kept)


def dedup_and_assign_ids() -> list[dict[str, Any]]:
    seen: dict[tuple, dict[str, Any]] = {}
    for rec in records:
        key = (rec["name"], rec["address"], rec["phone"])
        if key in seen:
            continue
        seen[key] = rec
    result = list(seen.values())
    counts: dict[str, int] = {}
    for rec in result:
        base = slug(f"hokkaido-{rec['establishment']}-{rec['school_type']}-{rec['municipality']}-{rec['name']}")
        counts[base] = counts.get(base, 0) + 1
        rec["id"] = base if counts[base] == 1 else f"{base}-{counts[base]}"
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=str(Path(__file__).resolve().parents[2] / "data" / "school-database" / "hokkaido.json"))
    args = parser.parse_args()

    parse_sorachi()
    parse_hidaka()
    parse_hiyama()
    parse_nemuro()
    parse_kushiro()
    parse_sapporo()
    parse_iburi()
    parse_ishikari()
    parse_shiribeshi()
    parse_soya()
    parse_okhotsk()
    parse_kamikawa()
    parse_tokachi()
    parse_oshima()
    parse_rumoi()
    parse_koukou_and_chuto()
    parse_tokubetsu_shien()
    parse_national()
    parse_private_excel()

    fill_missing_municipality()
    fix_misclassified_establishment()
    build_official_names()
    clean_records()
    drop_incomplete_records()
    fill_missing_postal()
    for rec in records:
        rec["name"] = collapse_repeated_block(rec["name"])
    result = dedup_and_assign_ids()
    result.sort(key=lambda r: (r["municipality"], r["school_type"], r["name"]))

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {len(result)} records to {out_path}")


if __name__ == "__main__":
    main()
