#!/usr/bin/env python3
"""埼玉県公式Excel名簿を学校検索用JSONへ変換する。

列番号ではなく見出し名から列を検出するため、校種ごとに多少レイアウトが
異なっても同じ処理で取り込める。原本は data-source/saitama/<year>/ に置く。
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


HEADER_ALIASES = {
    "school_name": ("学校名", "園名", "名称"),
    "school_name_kana": ("学校名ふりがな", "学校名かな", "ふりがな", "かな"),
    "municipality": ("市町村", "市区町村", "所在地市町村", "設置市町村"),
    "postal_code": ("郵便番号", "〒"),
    "address": ("住所", "所在地"),
    "phone": ("電話番号", "電話", "TEL", "Tel"),
    "operator": ("設置者", "学校法人", "法人名"),
    "course": ("課程", "課程等", "学部・課程"),
}

SCHOOL_TYPES = {
    "kindergarten": "幼稚園",
    "elementary": "小学校",
    "junior_high": "中学校",
    "compulsory": "義務教育学校",
    "high": "高等学校",
    "secondary": "中等教育学校",
    "special": "特別支援学校",
}


@dataclass(frozen=True)
class SourceSpec:
    path: Path
    school_type: str
    establishment_type: str
    source_name: str
    source_url: str
    source_date: str
    course: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return re.sub(r"\s+", " ", text)


def compact_header(value: Any) -> str:
    return re.sub(r"[\s　・･()（）\[\]【】]", "", normalize_text(value)).lower()


def normalize_postal_code(value: Any) -> str:
    digits = re.sub(r"\D", "", normalize_text(value))
    return f"{digits[:3]}-{digits[3:7]}" if len(digits) == 7 else normalize_text(value).replace("〒", "")


def normalize_phone(value: Any) -> str:
    text = normalize_text(value).replace("−", "-").replace("―", "-").replace("ー", "-")
    digits = re.sub(r"\D", "", text)
    if len(digits) == 10:
        if digits.startswith("048"):
            return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
        return text
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return text


def normalize_address(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("−", "-").replace("―", "-")
    if text and not text.startswith("埼玉県"):
        text = "埼玉県" + text
    return text


def infer_municipality(address: str) -> str:
    match = re.match(r"埼玉県(.+?(?:市|町|村))", address)
    return match.group(1) if match else ""


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).lower()
    normalized = re.sub(r"[^0-9a-zぁ-んァ-ヶ一-龠]+", "-", normalized).strip("-")
    return normalized or "school"


def find_header_row(rows: Iterable[tuple[Any, ...]], max_scan: int = 20) -> tuple[int, dict[str, int]]:
    best_row = -1
    best_map: dict[str, int] = {}
    for row_index, row in enumerate(rows, start=1):
        if row_index > max_scan:
            break
        current: dict[str, int] = {}
        for col_index, cell in enumerate(row):
            header = compact_header(cell)
            if not header:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if any(compact_header(alias) in header for alias in aliases):
                    current.setdefault(field, col_index)
        if "school_name" in current and "address" in current and len(current) > len(best_map):
            best_row, best_map = row_index, current
    if best_row < 0:
        raise ValueError("学校名・住所を含む見出し行を検出できませんでした")
    return best_row, best_map


def row_value(row: tuple[Any, ...], column_map: dict[str, int], field: str) -> str:
    index = column_map.get(field)
    return normalize_text(row[index]) if index is not None and index < len(row) else ""


def convert_source(spec: SourceSpec) -> list[dict[str, str]]:
    workbook = load_workbook(spec.path, data_only=True, read_only=True)
    records: list[dict[str, str]] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        try:
            header_row, column_map = find_header_row(rows)
        except ValueError:
            continue

        for row in rows[header_row:]:
            school_name = row_value(row, column_map, "school_name")
            address = normalize_address(row_value(row, column_map, "address"))
            if not school_name or not address:
                continue
            if any(word in school_name for word in ("計", "合計", "学校数")):
                continue

            municipality = row_value(row, column_map, "municipality") or infer_municipality(address)
            postal_code = normalize_postal_code(row_value(row, column_map, "postal_code"))
            phone = normalize_phone(row_value(row, column_map, "phone"))
            course = row_value(row, column_map, "course") or spec.course
            operator = row_value(row, column_map, "operator")

            stable_key = "|".join((spec.establishment_type, spec.school_type, municipality, school_name, course))
            record = {
                "id": f"saitama-{slug(stable_key)}",
                "prefecture": "埼玉県",
                "municipality": municipality,
                "school_name": school_name,
                "school_name_kana": row_value(row, column_map, "school_name_kana"),
                "school_type": spec.school_type,
                "establishment_type": spec.establishment_type,
                "operator": operator,
                "postal_code": postal_code,
                "address": address,
                "phone": phone,
                "course": course,
                "branch_type": "分校" if "分校" in school_name else "分教室" if "分教室" in school_name else "",
                "website": "",
                "source_name": spec.source_name,
                "source_url": spec.source_url,
                "source_date": spec.source_date,
                "verified_date": "",
            }
            records.append(record)
    return records


def load_manifest(path: Path, source_root: Path) -> list[SourceSpec]:
    data = json.loads(path.read_text(encoding="utf-8"))
    specs: list[SourceSpec] = []
    for item in data["sources"]:
        if not item.get("enabled", True):
            continue
        specs.append(
            SourceSpec(
                path=source_root / item["file"],
                school_type=item["school_type"],
                establishment_type=item["establishment_type"],
                source_name=item["source_name"],
                source_url=item["source_url"],
                source_date=item["source_date"],
                course=item.get("course", ""),
            )
        )
    return specs


def deduplicate(records: list[dict[str, str]]) -> list[dict[str, str]]:
    by_id: dict[str, dict[str, str]] = {}
    for record in records:
        record_id = record["id"]
        if record_id not in by_id:
            by_id[record_id] = record
            continue
        suffix = 2
        while f"{record_id}-{suffix}" in by_id:
            suffix += 1
        record["id"] = f"{record_id}-{suffix}"
        by_id[record["id"]] = record
    return list(by_id.values())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--manifest", type=Path, default=Path("tools/school-database/saitama_sources.json"))
    parser.add_argument("--source-root", type=Path, default=Path("data-source/saitama/2025"))
    parser.add_argument("--output", type=Path, default=Path("data/school-database/saitama.json"))
    args = parser.parse_args()

    all_records: list[dict[str, str]] = []
    for spec in load_manifest(args.manifest, args.source_root):
        if not spec.path.exists():
            print(f"SKIP missing: {spec.path}")
            continue
        converted = convert_source(spec)
        print(f"{spec.path.name}: {len(converted)} records")
        all_records.extend(converted)

    all_records = deduplicate(all_records)
    all_records.sort(key=lambda item: (item["school_type"], item["municipality"], item["school_name"], item["course"]))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(all_records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(all_records)} records to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
