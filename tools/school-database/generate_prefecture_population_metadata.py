#!/usr/bin/env python3
"""Generate 47-prefecture population metadata from the 2020 Census Table 2-1.

Source: 令和2年国勢調査 人口等基本集計 第2-1表
「男女，年齢（各歳），国籍総数か日本人別人口，平均年齢及び年齢中位数
－全国，都道府県，21大都市，特別区，人口50万以上の市」
統計表ID (statInfId): 000032142404
https://www.e-stat.go.jp/stat-search/files?cycle=0&tclass=000001125102

Only the 47 prefecture-level rows (地域識別コード = "a", excluding 全国),
国籍総数か日本人 = "1_うち日本人" (Japanese nationals), 男女 = "0_総数"
(both sexes) are used. Age "不詳" (unknown) rows are never apportioned
into single-year-age buckets; they are excluded from every sum, matching
the source file's own footnote convention.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE_FILE = ROOT / "data-source" / "census2020" / "table2-1.xlsx"
OUTPUT_PATH = ROOT / "data" / "school-database" / "prefecture-population.json"
PREFECTURE_METADATA_PATH = ROOT / "data" / "school-database" / "prefecture-card-metadata.json"

SHEET_NAME = "b02_01"
HEADER_ROW = 9
DATA_START_ROW = 12

TOTAL_COL = 5  # 000_総数
UNKNOWN_COL = 117  # 112_年齢「不詳」
MAX_SINGLE_AGE_COL = 116  # 110_109歳 .. 111_110歳以上 (age column = 6 + age)


def age_col(age: int) -> int:
    return 6 + age


def parse_cell_int(value: object) -> int:
    """Government tables use "-" for a suppressed/zero cell (seen here only
    at very old ages, e.g. 109/110歳以上, well outside the 3-17 range this
    script sums). Any other non-numeric value is treated as an error rather
    than silently coerced, since that could hide a real data problem."""
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str) and value.strip() == "-":
        return 0
    raise ValueError(f"unexpected non-numeric cell value: {value!r}")


AGE_GROUPS = {
    "census_preschool_3_5": {
        "label": "幼稚園相当年齢",
        "school_type": "幼稚園",
        "age_range": [3, 5],
        "ages": range(3, 6),
    },
    "census_elementary_6_11": {
        "label": "小学校相当年齢",
        "school_type": "小学校",
        "age_range": [6, 11],
        "ages": range(6, 12),
    },
    "census_junior_high_12_14": {
        "label": "中学校相当年齢",
        "school_type": "中学校",
        "age_range": [12, 14],
        "ages": range(12, 15),
    },
    "census_high_school_15_17": {
        "label": "高等学校相当年齢",
        "school_type": "高等学校",
        "age_range": [15, 17],
        "ages": range(15, 18),
    },
}

SOURCE_TITLE = (
    "令和2年国勢調査 人口等基本集計 第2-1表 男女，年齢（各歳），"
    "国籍総数か日本人別人口，平均年齢及び年齢中位数－全国，都道府県，"
    "21大都市，特別区，人口50万以上の市"
)
SOURCE_URL = "https://www.e-stat.go.jp/stat-search/files?cycle=0&tclass=000001125102"
SOURCE_TABLE_ID = "000032142404"
REFERENCE_DATE = "2020-10-01"
POPULATION_SCOPE = "census_japanese_population"
DEFINITION_NOTE = (
    "2020年10月1日現在の令和2年国勢調査による日本人人口（年齢各歳）。"
    "現在の人口ではない。年齢「不詳」の者は含まない。"
)

REGION_NAME_ROW = re.compile(r"^(\d{5})_(.+)$")


def share(numerator: int, denominator: int) -> float:
    return round((numerator / denominator) * 100, 6)


def load_prefecture_slug_map() -> dict[str, dict[str, str]]:
    """prefecture_name -> {prefecture_code, region_code} from the existing
    card metadata, which is the authoritative slug/region source for this
    site (avoids inventing a second naming scheme for the same 47 prefectures)."""
    payload = json.loads(PREFECTURE_METADATA_PATH.read_text(encoding="utf-8"))
    mapping = {}
    for pref in payload["prefectures"]:
        mapping[pref["prefecture_name"]] = {
            "prefecture_code": pref["prefecture_code"],
            "region_code": pref["region"]["code"],
        }
    return mapping


def extract_prefecture_rows(ws) -> list[dict[str, object]]:
    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        nationality = ws.cell(row=r, column=1).value
        sex = ws.cell(row=r, column=2).value
        region_code = ws.cell(row=r, column=3).value
        region_name_cell = ws.cell(row=r, column=4).value

        if nationality != "1_うち日本人" or sex != "0_総数" or region_code != "a":
            continue

        match = REGION_NAME_ROW.match(region_name_cell or "")
        if not match:
            raise ValueError(f"row {r}: unexpected region name {region_name_cell!r}")
        census_code, name = match.group(1), match.group(2)
        if census_code == "00000":
            continue  # 全国 (national total), not a prefecture

        total = parse_cell_int(ws.cell(row=r, column=TOTAL_COL).value)
        unknown = parse_cell_int(ws.cell(row=r, column=UNKNOWN_COL).value)
        age_values = {
            age: parse_cell_int(ws.cell(row=r, column=age_col(age)).value) for age in range(0, 111)
        }

        reconciled_total = sum(age_values.values()) + unknown
        if reconciled_total != total:
            raise ValueError(
                f"{name}: single-year ages + unknown ({reconciled_total}) != "
                f"published total ({total})"
            )

        rows.append(
            {
                "census_code": census_code,
                "name": name,
                "total": total,
                "unknown": unknown,
                "age_values": age_values,
            }
        )

    return rows


def build_payload(rows: list[dict[str, object]], slug_map: dict[str, dict[str, str]], accessed_at: str) -> dict[str, object]:
    if len(rows) != 47:
        raise ValueError(f"expected 47 prefecture rows, got {len(rows)}")

    prefectures = []
    for row in rows:
        name = row["name"]
        slug_info = slug_map.get(name)
        if slug_info is None:
            raise ValueError(f"no prefecture_code/region mapping found for census name {name!r}")

        total = row["total"]
        age_values = row["age_values"]
        age_3_17 = sum(age_values[age] for age in range(3, 18))

        age_groups = []
        for key, config in AGE_GROUPS.items():
            population = sum(age_values[age] for age in config["ages"])
            lo, hi = config["age_range"]
            age_groups.append(
                {
                    "key": key,
                    "label": config["label"],
                    "school_type": config["school_type"],
                    "age_range": config["age_range"],
                    "age_range_label": f"{lo}～{hi}歳",
                    "population": population,
                    "share_of_census_population_percent": share(population, total),
                }
            )

        prefectures.append(
            {
                "prefecture_code": slug_info["prefecture_code"],
                "prefecture_name": name,
                "census_prefecture_code": row["census_code"],
                "region_code": slug_info["region_code"],
                "population_scope": POPULATION_SCOPE,
                "reference_date": REFERENCE_DATE,
                "census_population": total,
                "census_age_3_17": age_3_17,
                "share_of_census_population_percent": share(age_3_17, total),
                "age_groups": age_groups,
                "source_title": SOURCE_TITLE,
                "source_url": SOURCE_URL,
                "source_table_id": SOURCE_TABLE_ID,
                "accessed_at": accessed_at,
                "definition_note": DEFINITION_NOTE,
            }
        )

    # Keep the same region-grouped order as the card metadata / portal display.
    name_order = list(slug_map.keys())
    prefectures.sort(key=lambda p: name_order.index(p["prefecture_name"]))

    return {
        "generated_at": accessed_at,
        "population_scope": POPULATION_SCOPE,
        "reference_date": REFERENCE_DATE,
        "source_title": SOURCE_TITLE,
        "source_url": SOURCE_URL,
        "source_table_id": SOURCE_TABLE_ID,
        "accessed_at": accessed_at,
        "definition_note": DEFINITION_NOTE,
        "prefectures": prefectures,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-file", type=Path, default=DEFAULT_SOURCE_FILE)
    parser.add_argument("--output", type=Path, default=OUTPUT_PATH)
    parser.add_argument("--accessed-at", default="2026-08-01")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.source_file, data_only=True)
    ws = wb[SHEET_NAME]
    rows = extract_prefecture_rows(ws)
    slug_map = load_prefecture_slug_map()
    payload = build_payload(rows, slug_map, args.accessed_at)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {args.output} ({len(payload['prefectures'])} prefectures)")


if __name__ == "__main__":
    main()
