#!/usr/bin/env python3
"""茨城県公式データを学校検索用JSONへ変換する。

原本は data-source/ibaraki/2025/ に置く（Git管理外）。

- public_school_list.xlsx: 茨城県教育委員会「市町村教育委員会・公立学校等一覧」
  （令和7年5月1日現在、複数シート）。**重要**: このExcelの「学校名」列は
  正式名称ではなく略称（例:「三の丸」「日立第一」「盲学校」）で記載されている。
  山梨県版で略称をそのまま収録してしまい利用者から訂正を受けた経緯があるため
  （docs/school-database/yamanashi/source-manifest.md 参照）、本スクリプトでは
  設置者（市町村名/「茨城県」/「国立大学法人茨城大学」）+ 略称 + 校種サフィックス
  を組み合わせて正式名称を構築する。構築ルールは複数の実例をWeb検索で裏取り
  済み（例:「水戸市立三の丸小学校」「茨城県立高萩清松高等学校」「茨城県立
  勝田中等教育学校」「日立市立日立特別支援学校」）。
- private_youchien.pdf: 茨城県福祉部子ども未来課「私立幼稚園一覧」（令和7年4月1日）。
  学校名は原本ですでに正式名称（省略なし）。
- private_elementary_junior_secondary.json / private_high_school.json:
  私立小学校・中学校・中等教育学校・高等学校一覧（茨城県教育委員会サイトの
  HTML表をそのまま書き起こしたもの。ダウンロード可能なxlsx/pdfが存在しないため）。
  学校名は原本ですでに正式名称。
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 正規化ユーティリティ
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\n", "").replace("\r", "")
    return text.strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"[ \t　]+", "", text)


def normalize_postal_code(value: Any) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    return normalize_text(value).replace("〒", "")


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    text = normalize_text(value)
    if text in ("", "-", "―", "ー", "‐", "－"):
        return ""
    return text.replace("−", "-").replace("―", "-").replace("ー", "-").replace("‐", "-").replace("－", "-")


def normalize_address(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"[ \t　]+", "", text)
    text = text.replace("−", "-").replace("―", "-")
    if text and not text.startswith("茨城県"):
        text = "茨城県" + text
    return text


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).lower()
    normalized = re.sub(r"[^0-9a-zぁ-んァ-ヶ一-龠]+", "-", normalized).strip("-")
    return normalized or "school"


# ---------------------------------------------------------------------------
# 正式名称の構築（山梨県版の反省を踏まえた最重要処理）
# ---------------------------------------------------------------------------

COMPLETE_NAME_SUFFIXES = (
    "幼稚園", "小学校", "中学校", "義務教育学校", "高等学校", "中等教育学校",
    "特別支援学校", "聾学校", "盲学校", "学園", "学院", "学部",
)


def build_official_name(*, committee_or_pref: str, short_name: str, suffix: str) -> str:
    """設置者 + 略称 + 校種サフィックスから正式名称を組み立てる。
    略称がすでに完全な名称（学校種を含む・「学園」等で終わる）場合は
    サフィックスを重ねて付けない。"""
    short_name = normalize_name(short_name)
    if any(short_name.endswith(s) for s in COMPLETE_NAME_SUFFIXES):
        core = short_name
    else:
        core = short_name + suffix

    if committee_or_pref == "茨城県":
        return f"茨城県立{core}"
    if committee_or_pref.startswith("国立大学法人"):
        # 国立学校名簿シートは原本にすでに正式名称が入っているため呼び出し元で
        # この関数自体を使わない想定だが、念のため素通しする。
        return core
    return f"{committee_or_pref}立{core}"


# ---------------------------------------------------------------------------
# 市町村表示順（市部 -> 郡部）
# ---------------------------------------------------------------------------

IBARAKI_CITIES = [
    "水戸市", "日立市", "土浦市", "古河市", "石岡市", "結城市", "龍ケ崎市", "下妻市",
    "常総市", "常陸太田市", "高萩市", "北茨城市", "笠間市", "取手市", "牛久市",
    "つくば市", "ひたちなか市", "鹿嶋市", "潮来市", "守谷市", "常陸大宮市", "那珂市",
    "筑西市", "坂東市", "稲敷市", "かすみがうら市", "桜川市", "神栖市", "行方市",
    "鉾田市", "つくばみらい市", "小美玉市",
]

IBARAKI_GUN_TOWNS = [
    "東茨城郡茨城町", "東茨城郡大洗町", "東茨城郡城里町",
    "那珂郡東海村",
    "久慈郡大子町",
    "稲敷郡美浦村", "稲敷郡阿見町", "稲敷郡河内町",
    "結城郡八千代町",
    "猿島郡五霞町", "猿島郡境町",
    "北相馬郡利根町",
]

MUNICIPALITY_ORDER = IBARAKI_CITIES + IBARAKI_GUN_TOWNS

_BARE_TOWN_TO_CANONICAL = {
    re.match(r"(東茨城郡|那珂郡|久慈郡|稲敷郡|結城郡|猿島郡|北相馬郡)(.+)", t).group(2): t
    for t in IBARAKI_GUN_TOWNS
}
_MUNICIPALITY_CANDIDATES = sorted(MUNICIPALITY_ORDER, key=len, reverse=True)
_BARE_TOWN_CANDIDATES = sorted(_BARE_TOWN_TO_CANONICAL, key=len, reverse=True)


def infer_municipality(address: str) -> str:
    text = address
    if text.startswith("茨城県"):
        text = text[len("茨城県"):]
    for candidate in _MUNICIPALITY_CANDIDATES:
        if text.startswith(candidate):
            return candidate
    for bare in _BARE_TOWN_CANDIDATES:
        if text.startswith(bare):
            return _BARE_TOWN_TO_CANONICAL[bare]
    return ""


SCHOOL_TYPE_ORDER = [
    "幼稚園", "小学校", "中学校", "義務教育学校",
    "高等学校", "中等教育学校", "特別支援学校",
]

WARNINGS: list[dict[str, str]] = []


def warn(context: str, message: str) -> None:
    WARNINGS.append({"context": context, "message": message})


def make_record(
    *, name: str, name_kana: str, postal_code: str, address: str, school_type: str,
    establishment: str, operator: str, phone: str, website: str,
    source_name: str, source_url: str, source_date: str, course: list[str],
) -> dict[str, Any]:
    name = normalize_name(name)
    address = normalize_address(address)
    municipality = infer_municipality(address)
    stable_key = "|".join((establishment, school_type, municipality, name, ",".join(course)))
    return {
        "id": f"ibaraki-{slug(stable_key)}",
        "prefecture": "茨城県",
        "name": name,
        "name_kana": normalize_name(name_kana),
        "postal_code": normalize_postal_code(postal_code),
        "address": address,
        "municipality": municipality,
        "school_type": school_type,
        "establishment": establishment,
        "operator": normalize_name(operator),
        "phone": normalize_phone(phone),
        "website": website,
        "source_name": source_name,
        "source_url": source_url,
        "source_date": source_date,
        "verified_date": "",
        "course": list(course),
    }


def split_course(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [c for c in re.split(r"[・\s　]+", text) if c]


# ---------------------------------------------------------------------------
# 公立Excel読み込み
# ---------------------------------------------------------------------------

PUBLIC_SOURCE_NAME = "市町村教育委員会・公立学校等一覧（茨城県教育委員会、令和7年5月1日現在）"
PUBLIC_SOURCE_URL = "https://kyoiku.pref.ibaraki.jp/about/survey/school-list/"
PUBLIC_SOURCE_DATE = "2025-05-01"

SKIP_NAME_WORDS = ("計", "合計")


def read_national(path: Path) -> list[dict[str, Any]]:
    """2-1国立（R7）: すでに正式名称。小学校/中学校がシート内で切り替わる。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb["2-1国立（R7）"]
    rows = list(sheet.iter_rows(values_only=True))
    records: list[dict[str, Any]] = []
    current_type = ""
    for row in rows:
        col0 = normalize_text(row[0]) if len(row) > 0 else ""
        if col0 in ("小学校", "中学校"):
            current_type = col0
            continue
        if col0 in ("学校名",):
            continue
        if not col0 or any(w in col0 for w in SKIP_NAME_WORDS):
            continue
        address = row[1] if len(row) > 1 else ""
        if not normalize_address(address):
            continue
        record = make_record(
            name=col0, name_kana="", postal_code=row[2] if len(row) > 2 else "",
            address=address, school_type=current_type, establishment="国立",
            operator="国立大学法人茨城大学", phone=row[3] if len(row) > 3 else "",
            website="", source_name=PUBLIC_SOURCE_NAME, source_url=PUBLIC_SOURCE_URL,
            source_date=PUBLIC_SOURCE_DATE, course=[],
        )
        records.append(record)
    return records


def read_prefectural_simple(
    path: Path, *, sheet_name: str, school_type: str, suffix: str, course_label: str = "",
) -> list[dict[str, Any]]:
    """県立中(附属)・県立中等: 1行1校（学級数等の統計列はあるが行の折返しは無い）。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    records: list[dict[str, Any]] = []
    for row in rows:
        col0 = normalize_text(row[0]) if len(row) > 0 else ""
        if col0 in ("学校名", "") or any(w in col0 for w in SKIP_NAME_WORDS):
            continue
        if "【" in col0 or col0 in ("中学校", "中等教育学校"):
            continue
        address = row[1] if len(row) > 1 else ""
        if not normalize_address(address):
            continue
        official_name = build_official_name(committee_or_pref="茨城県", short_name=col0, suffix=suffix)
        record = make_record(
            name=official_name, name_kana="", postal_code=row[2] if len(row) > 2 else "",
            address=address, school_type=school_type, establishment="公立",
            operator="茨城県", phone=row[3] if len(row) > 3 else "",
            website="", source_name=PUBLIC_SOURCE_NAME, source_url=PUBLIC_SOURCE_URL,
            source_date=PUBLIC_SOURCE_DATE, course=[course_label] if course_label else [],
        )
        records.append(record)
    return records


def read_prefectural_high_or_tokushi(
    path: Path, *, sheet_name: str, school_type: str, suffix: str, course_label: str = "",
) -> list[dict[str, Any]]:
    """県立高(全日/定時/通信)・県立特支: 学科/部別の内訳行が続くため、
    学校名が入っている行だけを本体行として扱い、以降の空白行は無視する。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    records: list[dict[str, Any]] = []
    for row in rows:
        col0 = normalize_text(row[0]) if len(row) > 0 else ""
        if not col0:
            continue  # 学科・部別の内訳継続行（学校名なし）
        if col0 in ("学校名",) or "【" in col0 or "高等学校・" in col0 or col0 == "特別支援学校":
            continue
        if any(w in col0 for w in SKIP_NAME_WORDS):
            continue
        address = row[1] if len(row) > 1 else ""
        if not normalize_address(address):
            continue
        official_name = build_official_name(committee_or_pref="茨城県", short_name=col0, suffix=suffix)
        record = make_record(
            name=official_name, name_kana="", postal_code=row[2] if len(row) > 2 else "",
            address=address, school_type=school_type, establishment="公立",
            operator="茨城県", phone=row[3] if len(row) > 3 else "",
            website="", source_name=PUBLIC_SOURCE_NAME, source_url=PUBLIC_SOURCE_URL,
            source_date=PUBLIC_SOURCE_DATE, course=[course_label] if course_label else [],
        )
        record["_merge_key"] = official_name
        records.append(record)
    return records


def merge_high_school_courses(all_course_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for record in all_course_records:
        groups.setdefault(record["_merge_key"], []).append(record)

    merged: list[dict[str, Any]] = []
    for key, group in groups.items():
        for r in group:
            r.pop("_merge_key", None)
        if len(group) == 1:
            merged.append(group[0])
            continue
        addresses = {g["address"] for g in group}
        if len(addresses) > 1:
            warn("high_school_course_merge", f"{key}: 課程間で所在地が異なるため統合せず個別レコードとして保持 (addresses={addresses})")
            merged.extend(group)
            continue
        base = dict(group[0])
        courses: list[str] = []
        for g in group:
            for c in g["course"]:
                if c not in courses:
                    courses.append(c)
        base["course"] = courses
        stable_key = "|".join((base["establishment"], base["school_type"], base["municipality"], base["name"], ",".join(base["course"])))
        base["id"] = f"ibaraki-{slug(stable_key)}"
        merged.append(base)
    return merged


def read_municipal(
    path: Path, *, sheet_name: str, school_type: str, suffix: str,
) -> list[dict[str, Any]]:
    """市町村立小/中/義務教育学校/特別支援学校: 市町村 | 学校名 | (よみがな) | 所在地 | 郵便番号 | 電話番号 | FAX番号"""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    records: list[dict[str, Any]] = []

    # 3-4市町村立特 は「所在地」列が3列目、他は4列目。ヘッダー行から列位置を検出する。
    header_row = None
    for row in rows:
        joined = "".join(normalize_text(c) for c in row if c)
        if "学校名" in joined:
            header_row = row
            break
    has_kana = header_row is not None and any(normalize_text(c) == "よみがな" for c in header_row)

    current_muni = ""
    for row in rows:
        col_muni = normalize_text(row[0]) if len(row) > 0 else ""
        col_name = normalize_text(row[1]) if len(row) > 1 else ""
        if col_muni in ("市町村",) or col_name in ("学校名",):
            continue
        if col_muni and "市町村立" not in col_muni:
            current_muni = col_muni
        if not col_name:
            continue

        if has_kana:
            kana = row[2] if len(row) > 2 else ""
            address = row[3] if len(row) > 3 else ""
            postal = row[4] if len(row) > 4 else ""
            phone = row[5] if len(row) > 5 else ""
        else:
            kana = ""
            address = row[2] if len(row) > 2 else ""
            postal = row[3] if len(row) > 3 else ""
            phone = row[4] if len(row) > 4 else ""

        address_text = normalize_text(address)
        if not address_text:
            continue
        # 市町村立小・中・義務教育学校の「所在地」列は市町村名を含まない
        # （例:「三の丸1-6-51」）ため、municipality判定のために市町村名を補う。
        # 3-4市町村立特のように既に市町村名を含む場合は重複させない。
        if current_muni and not address_text.startswith(current_muni):
            address_text = current_muni + address_text
        if not current_muni:
            warn("municipal", f"設置者（市町村）が特定できないため除外: {col_name}")
            continue

        official_name = build_official_name(committee_or_pref=current_muni, short_name=col_name, suffix=suffix)
        record = make_record(
            name=official_name, name_kana=kana, postal_code=postal, address=address_text,
            school_type=school_type, establishment="公立", operator=current_muni,
            phone=phone, website="", source_name=PUBLIC_SOURCE_NAME,
            source_url=PUBLIC_SOURCE_URL, source_date=PUBLIC_SOURCE_DATE, course=[],
        )
        records.append(record)
    return records


# ---------------------------------------------------------------------------
# 私立幼稚園PDF読み込み
# ---------------------------------------------------------------------------

PRIVATE_YOUCHIEN_SOURCE_NAME = "私立幼稚園一覧（茨城県福祉部子ども未来課、令和7年4月1日現在）"
PRIVATE_YOUCHIEN_SOURCE_URL = "http://www.kids.pref.ibaraki.jp/~kids/kosodate/nursing/nursing02/01/2504_youchien.pdf"


def read_private_youchien(path: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    cells = [normalize_text(c) for c in row]
                    if not cells or cells[0] in ("市町村", ""):
                        continue
                    if len(cells) < 4:
                        continue
                    name = cells[1]
                    address = cells[3]
                    if not name or not normalize_address(address):
                        continue
                    postal = cells[2] if len(cells) > 2 else ""
                    phone = cells[4] if len(cells) > 4 else ""
                    records.append(make_record(
                        name=name, name_kana="", postal_code=postal, address=address,
                        school_type="幼稚園", establishment="私立", operator="",
                        phone=phone, website="", source_name=PRIVATE_YOUCHIEN_SOURCE_NAME,
                        source_url=PRIVATE_YOUCHIEN_SOURCE_URL, source_date="2025-04-01", course=[],
                    ))
    return records


# ---------------------------------------------------------------------------
# 私立小中中等・高校（手書きJSON、原本にダウンロード可能ファイルなし）
# ---------------------------------------------------------------------------

def read_private_json_elementary_junior_secondary(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    source_name = data["source_name"]
    source_url = data["source_url"]
    records: list[dict[str, Any]] = []

    for item in data.get("elementary", []):
        records.append(make_record(
            name=item["name"], name_kana="", postal_code=item["postal_code"],
            address=item["address"], school_type="小学校", establishment="私立",
            operator=item["operator"], phone=item["phone"], website="",
            source_name=source_name, source_url=source_url, source_date="2026-07-27", course=[],
        ))
    for item in data.get("junior_high", []):
        records.append(make_record(
            name=item["name"], name_kana="", postal_code=item["postal_code"],
            address=item["address"], school_type="中学校", establishment="私立",
            operator=item["operator"], phone=item["phone"], website="",
            source_name=source_name, source_url=source_url, source_date="2026-07-27", course=[],
        ))
    for item in data.get("secondary", []):
        records.append(make_record(
            name=item["name"], name_kana="", postal_code=item["postal_code"],
            address=item["address"], school_type="中等教育学校", establishment="私立",
            operator=item["operator"], phone=item["phone"], website="",
            source_name=source_name, source_url=source_url, source_date="2026-07-27", course=[],
        ))
    return records


def read_private_json_high_school(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    source_name = data["source_name"]
    source_url = data["source_url"]

    course_records = []
    for item in data.get("schools", []):
        record = make_record(
            name=item["name"], name_kana="", postal_code=item["postal_code"],
            address=item["address"], school_type="高等学校", establishment="私立",
            operator=item["operator"], phone=item["phone"], website="",
            source_name=source_name, source_url=source_url, source_date="2026-07-27",
            course=[item["course"]],
        )
        record["_merge_key"] = record["name"]
        course_records.append(record)

    return merge_high_school_courses(course_records)


# ---------------------------------------------------------------------------
# メイン変換処理
# ---------------------------------------------------------------------------

def deduplicate(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id: dict[str, dict[str, Any]] = {}
    for record in records:
        record_id = record["id"]
        if record_id not in by_id:
            by_id[record_id] = record
            continue
        suffix = 2
        while f"{record_id}-{suffix}" in by_id:
            suffix += 1
        record["id"] = f"{record_id}-{suffix}"
        by_id[record["id"]] = record
    return list(by_id.values())


def sort_key(record: dict[str, Any]):
    m_idx = MUNICIPALITY_ORDER.index(record["municipality"]) if record["municipality"] in MUNICIPALITY_ORDER else 999
    t_idx = SCHOOL_TYPE_ORDER.index(record["school_type"]) if record["school_type"] in SCHOOL_TYPE_ORDER else 999
    return (m_idx, t_idx, record["name"])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, default=Path("data-source/ibaraki/2025"))
    parser.add_argument("--output", type=Path, default=Path("data/school-database/ibaraki.json"))
    parser.add_argument("--warnings-output", type=Path,
                         default=Path("tools/school-database/ibaraki_conversion_warnings.json"))
    args = parser.parse_args()

    public_path = args.source_root / "public_school_list.xlsx"
    all_records: list[dict[str, Any]] = []

    records = read_national(public_path)
    print(f"national: {len(records)} records")
    all_records.extend(records)

    records = read_prefectural_simple(
        public_path, sheet_name="2-2-1県立中（R7）", school_type="中学校", suffix="中学校")
    print(f"prefectural attached junior high: {len(records)} records")
    all_records.extend(records)

    records = read_prefectural_simple(
        public_path, sheet_name="2-2-2県立中等（R7）", school_type="中等教育学校", suffix="中等教育学校")
    print(f"prefectural secondary: {len(records)} records")
    all_records.extend(records)

    high_course_records = []
    high_course_records += read_prefectural_high_or_tokushi(
        public_path, sheet_name="2-2-3県立高全日制（R7）", school_type="高等学校",
        suffix="高等学校", course_label="全日制")
    high_course_records += read_prefectural_high_or_tokushi(
        public_path, sheet_name="2-2-4県立高定時制（R7）", school_type="高等学校",
        suffix="高等学校", course_label="定時制")
    high_course_records += read_prefectural_high_or_tokushi(
        public_path, sheet_name="2-2-5県立高通信制、専攻科（R7）", school_type="高等学校",
        suffix="高等学校", course_label="通信制")
    merged_high = merge_high_school_courses(high_course_records)
    print(f"prefectural high schools: {len(merged_high)} records ({len(high_course_records)} rows before merge)")
    all_records.extend(merged_high)

    records = read_prefectural_high_or_tokushi(
        public_path, sheet_name="2-2-6県立特支（R7）", school_type="特別支援学校", suffix="特別支援学校")
    print(f"prefectural special-needs: {len(records)} records")
    all_records.extend(records)

    records = read_municipal(public_path, sheet_name="3-1市町村立小（R7）", school_type="小学校", suffix="小学校")
    print(f"municipal elementary: {len(records)} records")
    all_records.extend(records)

    records = read_municipal(public_path, sheet_name="3-2市町村立中（R7）", school_type="中学校", suffix="中学校")
    print(f"municipal junior high: {len(records)} records")
    all_records.extend(records)

    records = read_municipal(public_path, sheet_name="3-3市町村立義（R７）", school_type="義務教育学校", suffix="義務教育学校")
    print(f"municipal compulsory: {len(records)} records")
    all_records.extend(records)

    records = read_municipal(public_path, sheet_name="3-4市町村立特（R7）", school_type="特別支援学校", suffix="特別支援学校")
    print(f"municipal special-needs: {len(records)} records")
    all_records.extend(records)

    youchien_path = args.source_root / "private_youchien.pdf"
    if youchien_path.exists():
        records = read_private_youchien(youchien_path)
        print(f"private kindergarten: {len(records)} records")
        all_records.extend(records)
    else:
        warn("private", f"SKIP missing: {youchien_path}")

    ejs_path = args.source_root / "private_elementary_junior_secondary.json"
    if ejs_path.exists():
        records = read_private_json_elementary_junior_secondary(ejs_path)
        print(f"private elementary/junior/secondary: {len(records)} records")
        all_records.extend(records)
    else:
        warn("private", f"SKIP missing: {ejs_path}")

    high_path = args.source_root / "private_high_school.json"
    if high_path.exists():
        records = read_private_json_high_school(high_path)
        print(f"private high school: {len(records)} records")
        all_records.extend(records)
    else:
        warn("private", f"SKIP missing: {high_path}")

    all_records = deduplicate(all_records)
    all_records.sort(key=sort_key)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(all_records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {len(all_records)} records to {args.output}")

    args.warnings_output.parent.mkdir(parents=True, exist_ok=True)
    args.warnings_output.write_text(json.dumps(WARNINGS, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(WARNINGS)} warnings to {args.warnings_output}")
    for item in WARNINGS:
        print(f"WARN [{item['context']}] {item['message']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
