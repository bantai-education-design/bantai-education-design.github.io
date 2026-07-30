#!/usr/bin/env python3
"""広島県の公式資料を学校検索用JSONへ変換する。

原本は data-source/hiroshima/ に置く（Git管理外）。

広島県教育委員会は、広島市・福山市を除く市町村立小学校・中学校の所在地一覧を
教育事務所（西部・西部教育事務所芸北支所・東部・北部）単位のHTMLページに
掲載しており、市町村名の見出し（h2）ごとに学校名・ふりがな・郵便番号・
所在地・電話番号のテーブルが続く構造になっている。福山市・広島市は別ページ。
義務教育学校・高等学校（県立・市立）は、それぞれ全県分が1ページにまとまった
HTMLテーブルとして掲載されている。

広島市立小学校・中学校のみ広島市公式サイトに別掲載されており、こちらは
郵便番号列が無いため、日本郵便のKEN_ALL（郵便番号データ）で住所から
郵便番号を補完する。

特別支援学校は住所付き一覧が存在しないため、各校公式サイトを個別に確認し
手作業で書き起こしたTSVを用いる。

私立学校（幼稚園・幼稚園型認定こども園・幼保連携型認定こども園・小学校・
中学校・高等学校）は広島県庁サイトからExcel形式でダウンロードできる。

国立学校（広島大学附属校）は大学公式サイトの記載を直接収録する。
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
from bs4 import BeautifulSoup


# ---------------------------------------------------------------------------
# 正規化ユーティリティ
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\r", "").replace("​", "")
    return text.strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("\n", "")
    return re.sub(r"[ \t　]+", "", text)


def normalize_kana(value: Any) -> str:
    return normalize_name(value)


def normalize_postal_code(value: Any) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    return ""


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    text = normalize_text(value)
    parts = [p.strip() for p in text.split("\n") if p.strip()]
    text = " / ".join(parts)
    text = text.replace("‐", "-").replace("−", "-").replace("―", "-").replace("－", "-")
    text = re.sub(r"^\((\d{2,5})\)\s*", r"\1-", text)
    if text in ("", "-", "ー", "―"):
        return ""
    return text


def normalize_address(value: Any, *, prefix: str = "") -> str:
    text = normalize_text(value)
    text = text.replace("\n", "")
    text = text.replace("‐", "-").replace("−", "-").replace("－", "-")
    text = re.sub(r"^〒?\d{3}-?\d{4}", "", text).strip()
    if prefix and text and prefix not in text:
        for i in range(len(prefix)):
            if text.startswith(prefix[i:]):
                text = prefix[:i] + text
                break
        else:
            text = prefix + text
    if text and not text.startswith("広島県"):
        text = "広島県" + text
    return text


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).lower()
    normalized = re.sub(r"[^0-9a-z぀-ヿ一-鿿]+", "-", normalized)
    return normalized.strip("-")


# 「廿日市市」のように市名自体の末尾が「市」で終わるため、市区町村を
# 非貪欲マッチで抽出すると最初の「市」で止まってしまう地名。
MUNI_DOUBLE_SUFFIX_NAMES = ("廿日市市",)

MUNI_FROM_ADDR_RE = re.compile(r"^広島県(\S+?[市区町村])")


def municipality_from_address(addr: str) -> str:
    for name in MUNI_DOUBLE_SUFFIX_NAMES:
        if addr.startswith("広島県" + name):
            return name
    m = MUNI_FROM_ADDR_RE.match(addr)
    return m.group(1) if m else ""


SOURCE_DATE = "2026-05-01"
PREF = "広島県"
DATA_DIR = Path(__file__).resolve().parents[2] / "data-source" / "hiroshima"

records: list[dict[str, Any]] = []


def add_record(
    *,
    name: str,
    name_kana: str = "",
    school_type: str,
    establishment: str,
    postal_code: str,
    address: str,
    municipality: str,
    phone: str,
    website: str = "",
    course: list[str] | None = None,
    source_name: str,
    source_url: str,
) -> None:
    name = normalize_name(name)
    if not name:
        return
    records.append({
        "id": "",
        "prefecture": PREF,
        "name": name,
        "name_kana": normalize_kana(name_kana),
        "postal_code": normalize_postal_code(postal_code) if postal_code else "",
        "address": address,
        "municipality": municipality,
        "school_type": school_type,
        "establishment": establishment,
        "operator": "",
        "phone": normalize_phone(phone),
        "website": website,
        "source_name": source_name,
        "source_url": source_url,
        "source_date": SOURCE_DATE,
        "verified_date": "",
        "course": course or [],
    })


# ---------------------------------------------------------------------------
# 学校名の正式名称化
# ---------------------------------------------------------------------------

FULL_NAME_SUFFIXES = (
    "幼稚園", "小学校", "中学校", "義務教育学校", "高等学校", "中等教育学校",
    "特別支援学校", "認定こども園", "こども園", "学園", "学院",
)

SINGLE_CHAR_SUFFIX = {"小学校": "小", "中学校": "中"}

INSTITUTION_PREFIX_RE = re.compile(r"^(広島市立|福山市立|市立|町立|村立|県立|国立)")

COURSE_QUALIFIER_RE = re.compile(r"[（(][^（）()]*[）)]\s*$")


def strip_course_qualifier(name: str) -> str:
    return COURSE_QUALIFIER_RE.sub("", name).strip()


def looks_like_official_name(name: str) -> bool:
    return "立" in name and any(name.endswith(s) for s in FULL_NAME_SUFFIXES)


def build_official_name(raw_name: str, school_type: str, establishment: str, municipality: str) -> str:
    """略称のまま収録された学校名を「{設置者}立{校名}{校種}」の正式名称に組み立てる。"""
    if establishment in ("私立", "国立"):
        return normalize_name(raw_name)

    name = strip_course_qualifier(normalize_name(raw_name))
    if looks_like_official_name(name):
        return name

    name = INSTITUTION_PREFIX_RE.sub("", name)
    if municipality and name.startswith(municipality + "立"):
        name = name[len(municipality) + 1:]
    if municipality and name.startswith(municipality):
        name = name[len(municipality):]

    full_suffix = school_type
    suffix = full_suffix
    if name.endswith(full_suffix):
        core = name[: -len(full_suffix)]
    elif name.endswith("分校") or name.endswith("分教室") or name.endswith("分級"):
        core = name
        suffix = ""
    else:
        single = SINGLE_CHAR_SUFFIX.get(school_type)
        if single and name.endswith(single):
            core = name[:-1]
        else:
            core = name

    prefix = municipality

    if not core:
        return f"{prefix}立{suffix}" if prefix else suffix

    return f"{prefix}立{core}{suffix}" if prefix else f"{core}{suffix}"


def collapse_repeated_block(name: str) -> str:
    n = len(name)
    for block_len in range(2, n // 2 + 1):
        if name[-block_len:] == name[-2 * block_len: -block_len]:
            return name[:-block_len]
    return name


# ---------------------------------------------------------------------------
# 郵便番号補完（KEN_ALLデータ、広島市立小・中学校の住所に使用）
# ---------------------------------------------------------------------------

_KEN_ALL_INDEX: dict[str, list[tuple[str, str]]] = {}


def _load_ken_all() -> None:
    if _KEN_ALL_INDEX:
        return
    path = DATA_DIR / "utf_ken_all.csv"
    if not path.exists():
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            cols = [c.strip('"') for c in line.rstrip("\n").split(",")]
            if len(cols) < 9 or cols[6] != "広島県":
                continue
            muni = cols[7]
            town = cols[8]
            postal = f"{cols[2][:3]}-{cols[2][3:]}"
            if town == "以下に掲載がない場合":
                town = ""
            _KEN_ALL_INDEX.setdefault(muni, []).append((town, postal))
    for muni in _KEN_ALL_INDEX:
        _KEN_ALL_INDEX[muni].sort(key=lambda t: -len(t[0]))


def postal_from_ken_all(municipality: str, address_after_muni: str) -> str:
    _load_ken_all()
    candidates = _KEN_ALL_INDEX.get(municipality)
    if not candidates:
        return ""
    for town, postal in candidates:
        if town and address_after_muni.startswith(town):
            return postal
    for town, postal in candidates:
        if not town:
            return postal
    return ""


def _known_municipality_prefix(addr_rest: str) -> str:
    """住所テキストの先頭が実在の広島県内市区町村名と一致すればそれを返す。"""
    _load_ken_all()
    for muni in _KEN_ALL_INDEX:
        if addr_rest.startswith(muni):
            return muni
    return ""


# ---------------------------------------------------------------------------
# 公立: 小学校・中学校（教育事務所別HTMLページ、市町村名は見出しh2）
# ---------------------------------------------------------------------------

def _parse_regional_page(path: Path, school_type: str, source_label: str) -> None:
    with open(path, encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")
    # 見出し(h2)が市町村ごとに続く構成（西部・東部等）と、単一市のみで見出しが
    # h1にしかない構成（福山市）の両方があるため、h2の直近の市町村名は
    # あくまでヒントとして使い、実際の市町村名は住所から都度確定させる。
    current_muni_hint = ""
    for el in soup.find_all(["h2", "table"]):
        if el.name == "h2":
            text = normalize_text(el.get_text())
            if re.search(r"[市区町村]$", text):
                current_muni_hint = text
            continue
        table = el
        rows = table.find_all("tr")
        if not rows:
            continue
        header = [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])]
        if not header or "学校名" not in header[0]:
            continue
        for tr in rows[1:]:
            cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) < 5:
                continue
            raw_name, kana, postal, addr, phone = cells[:5]
            name = normalize_name(raw_name)
            if not name:
                continue
            full_addr = normalize_address(addr, prefix=current_muni_hint)
            muni = municipality_from_address(full_addr) or current_muni_hint
            official = build_official_name(name, school_type, "公立", muni)
            add_record(
                name=official,
                name_kana=kana,
                school_type=school_type,
                establishment="公立",
                postal_code=postal,
                address=full_addr,
                municipality=muni,
                phone=phone,
                source_name=source_label,
                source_url="https://www.pref.hiroshima.lg.jp/site/kyouiku/fr-gakkou.html",
            )


REGIONAL_SHOU_FILES = [
    "pref_shou_seibu.html", "pref_shou_geihoku.html", "pref_shou_toubu.html",
    "pref_shou_hokubu.html", "pref_shou_fukuyama.html",
]
REGIONAL_CHU_FILES = [
    "pref_chu_seibu.html", "pref_chu_geihoku.html", "pref_chu_toubu.html",
    "pref_chu_hokubu.html", "pref_chu_fukuyama.html",
]


def parse_public_elem() -> None:
    for fname in REGIONAL_SHOU_FILES:
        _parse_regional_page(DATA_DIR / fname, "小学校", "広島県教育委員会 公立小学校一覧")


def parse_public_jhs() -> None:
    for fname in REGIONAL_CHU_FILES:
        _parse_regional_page(DATA_DIR / fname, "中学校", "広島県教育委員会 公立中学校一覧")


# ---------------------------------------------------------------------------
# 公立: 広島市立小学校・中学校（郵便番号なし、KEN_ALLで補完）
# ---------------------------------------------------------------------------

def _parse_hiroshima_city_page(path: Path, school_type: str) -> None:
    with open(path, encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")
    for h2 in soup.find_all("h2"):
        heading = normalize_text(h2.get_text())
        m = re.match(r"^(\S+?区)", heading)
        if not m:
            continue
        muni = "広島市" + m.group(1)
        table = h2.find_next("table")
        if not table:
            continue
        rows = table.find_all("tr")
        if not rows:
            continue
        header = [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])]
        if not header or "所在地" not in "".join(header):
            continue
        for tr in rows[1:]:
            cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) < 3:
                continue
            raw_name, addr, phone = cells[0], cells[1], cells[2]
            name = normalize_name(raw_name)
            if not name:
                continue
            full_addr = normalize_address(addr, prefix=muni)
            after_muni = full_addr[len("広島県" + muni):] if full_addr.startswith("広島県" + muni) else ""
            postal = postal_from_ken_all(muni, after_muni)
            official = build_official_name(name, school_type, "公立", "広島市")
            add_record(
                name=official,
                school_type=school_type,
                establishment="公立",
                postal_code=postal,
                address=full_addr,
                municipality=muni,
                phone=phone,
                source_name=f"広島市 市立{school_type}一覧",
                source_url="https://www.city.hiroshima.lg.jp/education/school/index.html",
            )


def parse_hiroshima_city_elem() -> None:
    _parse_hiroshima_city_page(DATA_DIR / "city_hiroshima_shou.html", "小学校")


def parse_hiroshima_city_jhs() -> None:
    _parse_hiroshima_city_page(DATA_DIR / "city_hiroshima_chu.html", "中学校")


# ---------------------------------------------------------------------------
# 公立: 義務教育学校（全県1ページ、市町村名は見出しh2、複数拠点セルに対応）
# ---------------------------------------------------------------------------

def parse_public_gimu() -> None:
    path = DATA_DIR / "pref_gimu.html"
    with open(path, encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")
    for h2 in soup.find_all("h2"):
        muni = normalize_text(h2.get_text())
        if not re.search(r"[市区町村]$", muni):
            continue
        table = h2.find_next("table")
        if not table:
            continue
        rows = table.find_all("tr")
        if not rows:
            continue
        header = [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])]
        if not header or "学校名" not in "".join(header):
            continue
        for tr in rows[1:]:
            cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) < 4:
                continue
            raw_name, postal, addr, phone = cells[0], cells[1], cells[2], cells[3]
            name = normalize_name(raw_name)
            if not name:
                continue
            postal_first = postal.split("\n")[0]
            addr_first = addr.split("\n")[0]
            full_addr = normalize_address(addr_first, prefix=muni)
            phone_parts = [normalize_phone(p) for p in phone.split("\n") if p.strip()]
            phone_val = " / ".join(dict.fromkeys(p for p in phone_parts if p))
            official = build_official_name(name, "義務教育学校", "公立", muni)
            add_record(
                name=official,
                school_type="義務教育学校",
                establishment="公立",
                postal_code=postal_first,
                address=full_addr,
                municipality=muni,
                phone=phone_val,
                source_name="広島県教育委員会 公立義務教育学校一覧",
                source_url="https://www.pref.hiroshima.lg.jp/site/kyouiku/16map-gimu-school-city.html",
            )


# ---------------------------------------------------------------------------
# 公立・市立: 高等学校（全日制・定時制・通信制、県立/市立混在の1ページ）
# ---------------------------------------------------------------------------

def parse_public_koukou() -> None:
    path = DATA_DIR / "pref_koukou_katei.html"
    with open(path, encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")
    tables = soup.find_all("table")
    course_labels = ["全日制", "定時制", "通信制"]
    for i, table in enumerate(tables[:3]):
        course = course_labels[i] if i < len(course_labels) else ""
        rows = table.find_all("tr")
        if not rows:
            continue
        header = [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])]
        if not header or "校名" not in "".join(header):
            continue
        current_est = "県立"
        for tr in rows[1:]:
            cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) == len(header) - 1:
                # 区分列はrowspanで省略されることがあるため先頭に空セルを補う
                cells = [""] + cells
            if len(cells) != len(header):
                continue
            row = dict(zip(header, cells))
            est_cell = row.get("区分", "").strip()
            if est_cell:
                current_est = "市立" if "市立" in est_cell else "県立"
            raw_name = normalize_name(row.get("校名", ""))
            raw_name = COURSE_QUALIFIER_RE.sub("", raw_name)
            if not raw_name:
                continue
            postal = row.get("郵便番号", "")
            addr = row.get("所在地", "")
            phone = row.get("電話番号", "")
            full_addr = normalize_address(addr)
            muni = municipality_from_address(full_addr)
            prefix_for_name = muni if current_est == "市立" else "広島県"
            official = build_official_name(raw_name, "高等学校", "公立", prefix_for_name)
            add_record(
                name=official,
                school_type="高等学校",
                establishment="公立",
                postal_code=postal,
                address=full_addr,
                municipality=muni,
                phone=phone,
                course=[course] if course else [],
                source_name="広島県教育委員会 県立・市立高等学校一覧",
                source_url="https://www.pref.hiroshima.lg.jp/site/kyouiku/14map-koukoumap-fr-katei.html",
            )


# ---------------------------------------------------------------------------
# 特別支援学校（各校公式サイトを個別確認して書き起こしたTSV）
# ---------------------------------------------------------------------------

def parse_tokushi() -> None:
    path = DATA_DIR / "tokushi_manual.tsv"
    lines = path.read_text(encoding="utf-8").splitlines()
    header = lines[0].split("\t")
    for line in lines[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        row = dict(zip(header, cols))
        addr = normalize_address(row.get("住所", ""))
        muni = municipality_from_address(addr)
        add_record(
            name=row.get("学校名", ""),
            school_type="特別支援学校",
            establishment="公立",
            postal_code=row.get("郵便番号", ""),
            address=addr,
            municipality=muni,
            phone=row.get("電話番号", ""),
            source_name="広島県内特別支援学校リンク集（各校公式サイト）",
            source_url="https://www.pref.hiroshima.lg.jp/site/kyouiku/14map-challenge-index.html",
        )


# ---------------------------------------------------------------------------
# 私立学校（Excel）
# ---------------------------------------------------------------------------

MUNI_SECTION_RE = re.compile(r"^[（(]\d+[）)]\s*(\S+?[市区町村])")


def _split_postal_addr(cell: str) -> tuple[str, str]:
    text = normalize_text(cell)
    # 全日制/通信制で校舎が別々の場合、所在地セルに複数の〒住所が
    # 改行区切りで入っていることがあるため、最初の住所のみを使う。
    first_block = re.split(r"\n(?=〒)", text)[0]
    m = re.match(r"〒?\s*(\d{3}-?\d{4})\s*(.*)$", first_block, re.S)
    if m:
        return normalize_postal_code(m.group(1)), m.group(2).replace("\n", "")
    return "", text.replace("\n", "")


def _parse_private_xlsx_simple(fname: str, school_type: str, source_label: str) -> None:
    """所在地列に市区町村名まで含む形式（私立小・中・高等学校）。"""
    wb = openpyxl.load_workbook(DATA_DIR / fname, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    for row in ws.iter_rows(values_only=False):
        values = [c.value for c in row]
        texts = [normalize_text(v) for v in values]
        if "学校名" in texts:
            header_row = texts
            continue
        if header_row is None:
            continue
        row_dict = dict(zip(header_row, values))
        name_cell = row_dict.get("学校名")
        if not name_cell:
            continue
        name = normalize_name(name_cell)
        addr_cell = row_dict.get("所在地")
        if not addr_cell:
            continue
        postal, addr_rest = _split_postal_addr(addr_cell)
        full_addr = normalize_address(addr_rest)
        muni = municipality_from_address(full_addr)
        phone = row_dict.get("電話", "")
        full_name = name if name.endswith(school_type) else name + school_type
        course = []
        katei = row_dict.get("課程")
        if katei:
            course = [normalize_text(katei)]
        add_record(
            name=full_name,
            school_type=school_type,
            establishment="私立",
            postal_code=postal,
            address=full_addr,
            municipality=muni,
            phone=phone,
            course=course,
            source_name=source_label,
            source_url="https://www.pref.hiroshima.lg.jp/site/hiroshimakennsiritugakkou/",
        )


def parse_private_elem() -> None:
    _parse_private_xlsx_simple("private_shou.xlsx", "小学校", "広島県 私立小学校名簿")


def parse_private_jhs() -> None:
    _parse_private_xlsx_simple("private_chu.xlsx", "中学校", "広島県 私立中学校名簿")


def parse_private_koukou() -> None:
    _parse_private_xlsx_simple("private_koukou.xlsx", "高等学校", "広島県 私立高等学校名簿")


def _parse_private_xlsx_sectioned(fname: str, school_type: str, source_label: str, name_suffix: str) -> None:
    """所在地列に市区町村名を含まず、（N）市区町村名 の見出し行で補う形式（幼稚園系）。"""
    wb = openpyxl.load_workbook(DATA_DIR / fname, data_only=True)
    ws = wb[wb.sheetnames[0]]
    current_muni = ""
    header_row = None
    for row in ws.iter_rows(values_only=False):
        values = [c.value for c in row]
        texts = [normalize_text(v) for v in values]
        joined = "".join(texts)
        sec_m = MUNI_SECTION_RE.match(joined)
        if sec_m:
            current_muni = sec_m.group(1)
            # 「廿日市市」のように市名自体が「市」で終わる地名は非貪欲マッチで
            # 途中の「市」で止まってしまうため、次の文字が「市」ならまとめて補う。
            if joined[sec_m.end(1):sec_m.end(1) + 1] == "市" and (current_muni + "市") in MUNI_DOUBLE_SUFFIX_NAMES:
                current_muni += "市"
            continue
        if "学校名" in texts:
            header_row = texts
            continue
        if header_row is None:
            continue
        row_dict = dict(zip(header_row, values))
        name_cell = row_dict.get("学校名")
        if not name_cell:
            continue
        name = normalize_name(name_cell)
        addr_cell = row_dict.get("所在地")
        if not addr_cell:
            continue
        postal, addr_rest = _split_postal_addr(addr_cell)
        # 所在地セルが既に（見出しの市区町村とは異なる）実在の市区町村名から
        # 始まっている場合（セクション見出しと実際の所在地が一致しない例外的な
        # ケース）は、見出し側の市区町村名を強引に連結せず、住所自身の記載を
        # 優先する。ただし「白島中町」のような単なる町名（市区町村名ではない）
        # と誤認しないよう、実在の広島県内市区町村名一覧と照合する。
        own_muni = _known_municipality_prefix(addr_rest)
        prefix_to_use = "" if own_muni else current_muni
        full_addr = normalize_address(addr_rest, prefix=prefix_to_use)
        muni = own_muni or current_muni
        phone = row_dict.get("電話", "")
        full_name = name if name.endswith(("幼稚園", "こども園")) else name + name_suffix
        add_record(
            name=full_name,
            school_type=school_type,
            establishment="私立",
            postal_code=postal,
            address=full_addr,
            municipality=muni,
            phone=phone,
            source_name=source_label,
            source_url="https://www.pref.hiroshima.lg.jp/site/hiroshimakennsiritugakkou/",
        )


def parse_private_youchien() -> None:
    _parse_private_xlsx_sectioned("private_youchien.xlsx", "幼稚園", "広島県 私立幼稚園名簿", "幼稚園")


def parse_private_youchien_ninka() -> None:
    _parse_private_xlsx_sectioned(
        "private_youchien_ninka.xlsx", "幼保連携型認定こども園",
        "広島県 私立幼稚園型認定こども園名簿", "こども園",
    )


def parse_private_youhoren() -> None:
    _parse_private_xlsx_sectioned(
        "private_youhoren.xlsx", "幼保連携型認定こども園",
        "広島県 私立幼保連携型認定こども園名簿", "こども園",
    )


# ---------------------------------------------------------------------------
# 国立学校: 広島大学附属校（大学公式サイト記載）
# ---------------------------------------------------------------------------

def parse_national() -> None:
    path = DATA_DIR / "national_schools.tsv"
    lines = path.read_text(encoding="utf-8").splitlines()
    header = lines[0].split("\t")
    for line in lines[1:]:
        if not line.strip():
            continue
        cols = line.split("\t")
        row = dict(zip(header, cols))
        addr = normalize_address(row.get("住所", ""))
        muni = municipality_from_address(addr)
        add_record(
            name=row.get("学校名", ""),
            school_type=row.get("学校種", ""),
            establishment="国立",
            postal_code=row.get("郵便番号", ""),
            address=addr,
            municipality=muni,
            phone=row.get("電話番号", ""),
            source_name="広島大学 附属学校園",
            source_url="https://www.hiroshima-u.ac.jp/fuzoku",
        )


# ---------------------------------------------------------------------------
# クリーニング・重複除去
# ---------------------------------------------------------------------------

ALLOWED_SCHOOL_TYPES = {
    "幼稚園", "幼保連携型認定こども園", "小学校", "中学校", "義務教育学校",
    "高等学校", "中等教育学校", "特別支援学校",
}

BARE_TYPE_NAMES = {"幼稚園", "小学校", "中学校", "高等学校", "特別支援学校"}


def clean_records() -> None:
    cleaned = []
    for rec in records:
        if rec["school_type"] not in ALLOWED_SCHOOL_TYPES:
            continue
        if rec["establishment"] not in ("公立", "私立", "国立"):
            continue
        if rec["name"] in BARE_TYPE_NAMES:
            continue
        if not rec["postal_code"]:
            continue
        if len(rec["address"]) < 6:
            continue
        cleaned.append(rec)
    records.clear()
    records.extend(cleaned)
    for rec in records:
        rec["name"] = collapse_repeated_block(rec["name"])


def dedup_and_assign_ids() -> list[dict[str, Any]]:
    seen: dict[tuple, dict[str, Any]] = {}
    for rec in records:
        key = (rec["name"], rec["address"], rec["phone"])
        if key in seen:
            continue
        seen[key] = rec
    result = list(seen.values())
    counts: dict[str, int] = {}
    for rec in result:
        base = slug(f"hiroshima-{rec['establishment']}-{rec['school_type']}-{rec['municipality']}-{rec['name']}")
        counts[base] = counts.get(base, 0) + 1
        rec["id"] = base if counts[base] == 1 else f"{base}-{counts[base]}"
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=str(Path(__file__).resolve().parents[2] / "data" / "school-database" / "hiroshima.json"))
    args = parser.parse_args()

    parse_public_elem()
    parse_public_jhs()
    parse_hiroshima_city_elem()
    parse_hiroshima_city_jhs()
    parse_public_gimu()
    parse_public_koukou()
    parse_tokushi()
    parse_private_elem()
    parse_private_jhs()
    parse_private_koukou()
    parse_private_youchien()
    parse_private_youchien_ninka()
    parse_private_youhoren()
    parse_national()

    clean_records()
    result = dedup_and_assign_ids()
    result.sort(key=lambda r: (r["municipality"], r["school_type"], r["name"]))

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"wrote {len(result)} records to {out_path}")


if __name__ == "__main__":
    main()
