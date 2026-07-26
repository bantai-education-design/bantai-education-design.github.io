#!/usr/bin/env python3
"""新潟県公式データを学校検索用JSONへ変換する。

原本は data-source/niigata/2025/ に置く（Git管理外）。

新潟県教育委員会が公開するExcel（市町村立小学校・中学校・義務教育学校、
県立中学校・中等教育学校、公立特別支援学校）は「新潟市立」を含まないため、
政令指定都市である新潟市が別途公開する学校便覧Excel（小学校・中学校・幼稚園の部）
を組み合わせて県内全体を網羅する。県立高等学校・中等教育学校・県立中学校は
所在地・電話番号のみを収録した概覧PDFから抽出し、全日制・定時制・通信制などで
同一校・同一住所の行は1校へ統合してcourseへ保持する（分校・分教室・キャンパスは
住所または電話番号が異なるため独立レコードとして保持）。

私立中学・高等学校および私立幼稚園は新潟県私立学校一覧PDFに基づくが、
テキスト抽出がテーブルの列単位（学校名列→課程列→学科列→住所/電話/HP列）に
なるため、原本を目視で照合したうえで本スクリプト内に直接値を保持する
（4校以下の新潟市立高等学校・中等教育学校・特別支援学校の分教科構成も同様）。
私立幼稚園一覧PDFには郵便番号が含まれないため、日本郵便の郵便番号データ
（住所の郵便番号、UTF-8版）で町域名から一意に特定できたものだけを補う。
専修学校・各種学校・認定こども園・大学は他県版と同様に初版対象外とする。
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# 正規化ユーティリティ（福島県版・宮城県版と共通の設計）
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\n", "").replace("\r", "")
    return text.strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"[ \t　]+", "", text)


def normalize_postal_code(value: Any) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    return normalize_text(value).replace("〒", "")


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    text = normalize_text(value)
    if text in ("", "-", "―", "ー"):
        return ""
    text = text.replace("−", "-").replace("―", "-").replace("ー", "-")
    text = re.sub(r"[（(]代[）)]|㈹", "", text)
    text = re.split(r"[・/]", text)[0].strip()
    m = re.match(r"^(0\d{1,4})\((\d{1,4})\)(\d{3,4})$", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return re.sub(r"[^\d-]", "", text)


def normalize_address(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("−", "-").replace("―", "-")
    # 新潟市立学校便覧の原本にある誤記（荻川小学校: 同区内の他11校は全て
    # 「秋葉区」と表記されているが、この1校のみ「区」が欠落している）を補う。
    text = text.replace("新潟市秋葉車場", "新潟市秋葉区車場")
    if text and not text.startswith("新潟県"):
        text = "新潟県" + text
    return text


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).lower()
    normalized = re.sub(r"[^0-9a-zぁ-んァ-ヶ一-龠]+", "-", normalized).strip("-")
    return normalized or "school"


# ---------------------------------------------------------------------------
# 市町村表示順（新潟市の区 -> 上越教育事務所管内 -> 中越教育事務所管内 ->
# 下越教育事務所管内、新潟県教育委員会公開資料の地域区分に準拠）
# ---------------------------------------------------------------------------

NIIGATA_CITY_WARDS = [
    "新潟市北区", "新潟市東区", "新潟市中央区", "新潟市江南区",
    "新潟市秋葉区", "新潟市南区", "新潟市西区", "新潟市西蒲区",
]
JOETSU = ["上越市", "糸魚川市", "妙高市"]
CHUETSU = [
    "長岡市", "三条市", "柏崎市", "小千谷市", "加茂市", "十日町市", "見附市",
    "燕市", "魚沼市", "南魚沼市", "弥彦村", "田上町", "出雲崎町", "湯沢町",
    "津南町", "刈羽村",
]
KAETSU = [
    "新発田市", "村上市", "五泉市", "佐渡市", "阿賀野市", "胎内市",
    "聖籠町", "阿賀町", "関川村", "粟島浦村",
]

MUNICIPALITY_ORDER = NIIGATA_CITY_WARDS + JOETSU + CHUETSU + KAETSU
_MUNICIPALITY_CANDIDATES = sorted(MUNICIPALITY_ORDER, key=len, reverse=True)

# 郡部の町村は住所表記が「郡名+町村名」になる場合があるため、郡名付き表記
# からも同じ市町村名へ引き当てられるようにする。
_DISTRICT_PREFIXED = {
    "阿賀町": "東蒲原郡阿賀町",
    "津南町": "中魚沼郡津南町",
    "聖籠町": "北蒲原郡聖籠町",
    "出雲崎町": "三島郡出雲崎町",
    "刈羽村": "刈羽郡刈羽村",
    "弥彦村": "西蒲原郡弥彦村",
    "湯沢町": "南魚沼郡湯沢町",
    "田上町": "南蒲原郡田上町",
    "粟島浦村": "岩船郡粟島浦村",
    "関川村": "岩船郡関川村",
}
_DISTRICT_CANDIDATES = sorted(_DISTRICT_PREFIXED.values(), key=len, reverse=True)

SCHOOL_TYPE_ORDER = [
    "幼稚園", "小学校", "中学校", "義務教育学校",
    "高等学校", "中等教育学校", "特別支援学校",
]


def infer_municipality(address: str) -> str:
    text = address
    if text.startswith("新潟県"):
        text = text[len("新潟県"):]
    for candidate in _MUNICIPALITY_CANDIDATES:
        if text.startswith(candidate):
            return candidate
    for district_form in _DISTRICT_CANDIDATES:
        if text.startswith(district_form):
            bare = next(k for k, v in _DISTRICT_PREFIXED.items() if v == district_form)
            return bare
    return ""


class WarningLog:
    def __init__(self) -> None:
        self.items: list[dict[str, str]] = []

    def add(self, context: str, message: str) -> None:
        self.items.append({"context": context, "message": message})


WARNINGS = WarningLog()


def make_record(
    *, name: str, name_kana: str = "", postal_code: str, address: str, school_type: str,
    establishment: str, operator: str, phone: str, website: str,
    source_name: str, source_url: str, source_date: str,
    course: list[str] | None = None,
) -> dict[str, Any]:
    name = normalize_name(name)
    address = normalize_address(address)
    municipality = infer_municipality(address)
    course = course or []
    stable_key = "|".join((establishment, school_type, municipality, name, ",".join(course)))
    return {
        "id": f"niigata-{slug(stable_key)}",
        "prefecture": "新潟県",
        "name": name,
        "name_kana": normalize_name(name_kana),
        "postal_code": normalize_postal_code(postal_code),
        "address": address,
        "municipality": municipality,
        "school_type": school_type,
        "establishment": establishment,
        "operator": normalize_name(operator),
        "phone": normalize_phone(phone),
        "website": website,
        "source_name": source_name,
        "source_url": source_url,
        "source_date": source_date,
        "verified_date": "",
        "course": list(course),
    }


# ---------------------------------------------------------------------------
# 新潟県公開Excel（市町村立小学校・中学校、義務教育学校、県立中学校・中等
# 教育学校）: 表構造は共通で [№, (地域, 地域No.,) 市町村, 学校名, 郵便番号,
# 所在地, 電話番号]
# ---------------------------------------------------------------------------

PREF_SOURCE_DATE = "2026-04-01"
PREF_SOURCE_URL = "https://www.pref.niigata.lg.jp/site/kyoiku/1196007390231.html"


def read_pref_simple_excel(
    path: Path, sheet_name: str, *, school_type: str, source_name: str,
    has_region_columns: bool,
) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None or not isinstance(row[0], (int, float)):
            continue
        if has_region_columns:
            municipality, name, postal, address, phone = row[3], row[4], row[5], row[6], row[7]
        else:
            municipality, name, postal, address, phone = row[1], row[2], row[3], row[4], row[5]
        name = normalize_name(name)
        if not name:
            continue
        establishment = "公立"
        operator = "新潟県" if normalize_text(municipality) == "県立" else normalize_text(municipality)
        row_school_type = school_type
        if school_type == "中等教育学校" and name.endswith("中学校") and not name.endswith("中等教育学校"):
            row_school_type = "中学校"
        records.append(make_record(
            name=name,
            postal_code=postal,
            address=address,
            school_type=row_school_type,
            establishment=establishment,
            operator=operator,
            phone=phone,
            website="",
            source_name=source_name,
            source_url=PREF_SOURCE_URL,
            source_date=PREF_SOURCE_DATE,
        ))
    return records


def read_pref_special_needs_excel(path: Path) -> list[dict[str, Any]]:
    """公立特別支援学校（県立・市立混在、市町村欄なし。住所から市町村を推定）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["特別支援"]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None or not isinstance(row[0], (int, float)):
            continue
        name, postal, address, phone = row[1], row[2], row[3], row[4]
        name = normalize_name(name)
        if not name:
            continue
        operator = "新潟県" if name.startswith("県立") else re.match(r"^(\S+?[市町村])立", name)
        if isinstance(operator, re.Match):
            operator = operator.group(1)
        elif operator != "新潟県":
            operator = ""
        records.append(make_record(
            name=name,
            postal_code=postal,
            address=address,
            school_type="特別支援学校",
            establishment="公立",
            operator=operator,
            phone=phone,
            website="",
            source_name="新潟県公立特別支援学校一覧（令和8年度）",
            source_url=PREF_SOURCE_URL,
            source_date=PREF_SOURCE_DATE,
        ))
    return records


# ---------------------------------------------------------------------------
# 新潟市立学校便覧（小学校・中学校）: 1校1行だが郵便番号・所在地が隣接2列に
# 分割されている（例: col16="950-", col17="3126"）。
# ---------------------------------------------------------------------------

CITY_SOURCE_DATE = "2026-05-01"
CITY_SOURCE_URL = "https://www.city.niigata.lg.jp/kosodate/gakko/gakko_keikaku/indexbinran.html"


def read_city_two_col_excel(
    path: Path, sheet_name: str, *, school_type: str, source_name: str,
) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not isinstance(row[0], (int, float)):
            continue
        name = normalize_name(row[2])
        if not name:
            continue
        postal = f"{normalize_text(row[16])}{normalize_text(row[17])}"
        address = f"{normalize_text(row[18])}{normalize_text(row[19])}"
        phone = row[20]
        records.append(make_record(
            name=name,
            postal_code=postal,
            address=address,
            school_type=school_type,
            establishment="公立",
            operator="新潟市",
            phone=phone,
            website="",
            source_name=source_name,
            source_url=CITY_SOURCE_URL,
            source_date=CITY_SOURCE_DATE,
        ))
    return records


def read_city_kindergarten_excel(path: Path) -> list[dict[str, Any]]:
    """新潟市立幼稚園: 1園が2物理行にまたがり、郵便番号・所在地は同じ列で
    1行目に前半、2行目に後半が入る（電話は1行目のみ）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["幼稚園"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    records: list[dict[str, Any]] = []
    i = 0
    while i < len(rows):
        row = rows[i]
        if row and isinstance(row[0], (int, float)) and row[2]:
            name = normalize_name(row[2])
            postal = normalize_text(row[21])
            address = normalize_text(row[22])
            phone = row[23]
            if i + 1 < len(rows):
                nxt = rows[i + 1]
                if not (nxt[0] and isinstance(nxt[0], (int, float))):
                    postal += normalize_text(nxt[21])
                    address += normalize_text(nxt[22])
            records.append(make_record(
                name=name,
                postal_code=postal,
                address=address,
                school_type="幼稚園",
                establishment="公立",
                operator="新潟市",
                phone=phone,
                website="",
                source_name="令和8年度新潟市立学校一覧 幼稚園の部",
                source_url=CITY_SOURCE_URL,
                source_date=CITY_SOURCE_DATE,
            ))
        i += 1
    return records


# ---------------------------------------------------------------------------
# 新潟市立 高等学校・中等教育学校・特別支援学校（各1〜2校のみ。学部・課程
# ごとに行が分かれる結合セル構造のため原本を目視で確認しスクリプトに直接
# 保持する。新潟市学校便覧 令和8年度版より）。
# ---------------------------------------------------------------------------

NIIGATA_CITY_SMALL_SCHOOLS: list[dict[str, Any]] = [
    dict(name="万代高等学校", postal_code="950-0075",
         address="新潟市中央区沼垂東6丁目8番1号", phone="025-241-0193",
         school_type="高等学校", course=["全日制"]),
    dict(name="明鏡高等学校", postal_code="950-0075",
         address="新潟市中央区沼垂東6丁目11番1号", phone="025-246-3535",
         school_type="高等学校", course=["定時制"]),
    dict(name="高志中等教育学校", postal_code="950-0926",
         address="新潟市中央区高志1丁目15番1号", phone="025-286-9811",
         school_type="中等教育学校", course=[]),
    dict(name="東特別支援学校", postal_code="950-0806",
         address="新潟市東区海老ケ瀬31番地", phone="025-271-9117",
         school_type="特別支援学校", course=[]),
    dict(name="西特別支援学校", postal_code="953-0043",
         address="新潟市西蒲区堀山新田88番地", phone="0256-73-3311",
         school_type="特別支援学校", course=[]),
]


def build_niigata_city_small_schools() -> list[dict[str, Any]]:
    return [
        make_record(
            name=item["name"], postal_code=item["postal_code"], address=item["address"],
            school_type=item["school_type"], establishment="公立", operator="新潟市",
            phone=item["phone"], website="",
            source_name=f"令和8年度新潟市立学校一覧 {item['school_type']}の部",
            source_url=CITY_SOURCE_URL, source_date=CITY_SOURCE_DATE,
            course=item["course"],
        )
        for item in NIIGATA_CITY_SMALL_SCHOOLS
    ]


# ---------------------------------------------------------------------------
# 県立高等学校・中等教育学校・県立中学校（概覧PDFのテキスト抽出結果を解析）。
# PDFの表は「学校名＋所在地」列が行順に並んだ後、「電話（代表）」列が
# TEL/FAXの2行1組でまとめて後続する（PDFの列単位抽出による）。同一名・
# 同一住所・同一電話番号の行は課程違い（全日制/定時制/通信制）として1校へ
# 統合する。
# ---------------------------------------------------------------------------

GAIRAN_SOURCE_NAME = "新潟県立高等学校・中等教育学校 学校概覧（所在地、電話・FAX番号、令和8年度）"
GAIRAN_SOURCE_URL = "https://www.pref.niigata.lg.jp/uploaded/life/824462_2656520_misc.pdf"
GAIRAN_SOURCE_DATE = "2026-04-01"

SECTION_HEADER_RE = re.compile(r"^■(県立高等学校|県立中等教育学校|県立中学校)\s*(全日制課程|定時制課程|通信制課程)?$")
NAME_LINE_RE = re.compile(r"^(?:(\d+)\s+)?(.+?(?:高等学校|中等教育学校|中学校)(?:\S*キャンパス|\S*分校)?)(?:\s+(.+))?$")
ADDRESS_LINE_RE = re.compile(r"^(\d{3}-\d{4})[　 ](.+)$")
TEL_RE = re.compile(r"^\s*TEL\s*([\d-]+)$")
FAX_RE = re.compile(r"^\s*FAX\s*([\d-]+)$")


def read_gairan_pdf_text(path: Path) -> list[dict[str, Any]]:
    raw_lines = [normalize_text(line) for line in path.read_text(encoding="utf-8").splitlines()]
    page_title = "新潟県立高等学校・中等教育学校 学校概覧(所在地、電話・FAX番号)"
    lines = [line for line in raw_lines if line and line != page_title]

    school_type_map = {
        "県立高等学校": "高等学校",
        "県立中等教育学校": "中等教育学校",
        "県立中学校": "中学校",
    }

    # PDFの列単位抽出により、学校名＋所在地の行が複数の見出し（課程・学校種）
    # にまたがって連続したあとに、電話ブロックがまとまって複数回出現する
    # ことがある（県立中等教育学校のページなど）。そのため学校名＋所在地は
    # 見出しをまたいでフラットな1本のリストへ追記し、電話ブロックはその都度
    # 「まだ電話番号が入っていない先頭からブロックの件数分」に割り当てる
    # カーソル方式にする。
    entries: list[dict[str, Any]] = []
    phone_cursor = 0
    i = 0
    n = len(lines)
    current_school_type = None
    current_course_label = None

    while i < n:
        line = lines[i]
        header_m = SECTION_HEADER_RE.match(line)
        if header_m:
            current_school_type = school_type_map[header_m.group(1)]
            current_course_label = header_m.group(2)
            if current_course_label:
                current_course_label = current_course_label.replace("課程", "")
            i += 1
            # skip "学校名" / "所在地" column headers
            while i < n and lines[i] in ("学校名", "所在地"):
                i += 1
            continue

        if line == "電話(代表)":
            i += 1
            phone_lines = []
            while i < n and (TEL_RE.match(lines[i]) or FAX_RE.match(lines[i])):
                phone_lines.append(lines[i])
                i += 1
            pairs = []
            j = 0
            while j < len(phone_lines):
                tel_m = TEL_RE.match(phone_lines[j])
                fax_m = FAX_RE.match(phone_lines[j + 1]) if j + 1 < len(phone_lines) else None
                if tel_m:
                    pairs.append(tel_m.group(1))
                    j += 2 if fax_m else 1
                else:
                    j += 1
            for tel in pairs:
                if phone_cursor < len(entries):
                    entries[phone_cursor]["phone"] = tel
                    phone_cursor += 1
            continue

        name_m = NAME_LINE_RE.match(line)
        if name_m and current_school_type:
            name = name_m.group(2).strip()
            inline_address = name_m.group(3)
            if inline_address:
                addr_m = ADDRESS_LINE_RE.match(inline_address.strip())
                postal, address = (addr_m.group(1), addr_m.group(2)) if addr_m else ("", inline_address)
                i += 1
            else:
                i += 1
                addr_m = ADDRESS_LINE_RE.match(lines[i]) if i < n else None
                if addr_m:
                    postal, address = addr_m.group(1), addr_m.group(2)
                    i += 1
                else:
                    postal, address = "", ""
            entries.append({
                "name": name,
                "postal_code": postal,
                "address": address,
                "school_type": current_school_type,
                "course": [current_course_label] if current_course_label else [],
                "phone": "",
            })
            continue

        i += 1

    return entries


def _merge_address_key(address: str) -> str:
    """課程違いの行では「番地」の有無だけが異なる同一住所表記があるため、
    統合キー算出時のみ末尾の「番地」を取り除いて比較する（保存値は変更しない）。"""
    return address[:-2] if address.endswith("番地") else address


def merge_gairan_entries(entries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """同一名・同一住所は1校へ統合し、courseを合算する（電話番号は課程間で
    異なる場合があるため、統合キーには使用しない。最初に出現した番号を残す）。"""
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    order: list[tuple[str, str]] = []
    for entry in entries:
        key = (entry["name"], _merge_address_key(entry["address"]))
        if key not in merged:
            merged[key] = dict(entry)
            merged[key]["course"] = list(entry["course"])
            order.append(key)
        else:
            existing = merged[key]
            for c in entry["course"]:
                if c not in existing["course"]:
                    existing["course"].append(c)
            if not existing["phone"]:
                existing["phone"] = entry["phone"]
    return [merged[key] for key in order]


def build_pref_high_schools() -> list[dict[str, Any]]:
    """概覧PDFからは高等学校のみを採用する（中等教育学校・県立中学校は
    pref_middle_secondary.xlsx の方が構造化されており重複するため）。"""
    path = SRC_ROOT / "gakkou_gairan.txt"
    entries = read_gairan_pdf_text(path)
    entries = [e for e in entries if e["school_type"] == "高等学校"]
    entries = merge_gairan_entries(entries)
    records = []
    for entry in entries:
        records.append(make_record(
            name=entry["name"],
            postal_code=entry["postal_code"],
            address=entry["address"],
            school_type=entry["school_type"],
            establishment="公立",
            operator="新潟県",
            phone=entry["phone"],
            website="",
            source_name=GAIRAN_SOURCE_NAME,
            source_url=GAIRAN_SOURCE_URL,
            source_date=GAIRAN_SOURCE_DATE,
            course=entry["course"],
        ))
    return records


# ---------------------------------------------------------------------------
# 私立中学校・高等学校（一覧PDFのテキスト抽出は列単位のため、原本を目視で
# 照合した確定値をここに保持する。令和8年6月3日更新版より）。
# ---------------------------------------------------------------------------

PRIVATE_SOURCE_DATE = "2026-06-03"
PRIVATE_HS_SOURCE_URL = "https://www.pref.niigata.lg.jp/uploaded/attachment/489726.pdf"
PRIVATE_KG_SOURCE_URL = "https://www.pref.niigata.lg.jp/uploaded/attachment/498473.pdf"

PRIVATE_HIGH_SCHOOLS: list[dict[str, Any]] = [
    dict(name="新潟明訓高等学校", course=["全日制"], postal_code="950-0116",
         address="新潟市江南区北山1037番地", phone="025-257-2131",
         website="http://www.niigata-meikun.ed.jp/"),
    dict(name="北越高等学校", course=["全日制"], postal_code="950-0916",
         address="新潟市中央区米山5丁目12番1号", phone="025-245-5681",
         website="http://www.hokuetsu.ed.jp/"),
    dict(name="新潟青陵高等学校", course=["全日制", "通信制"], postal_code="951-8121",
         address="新潟市中央区水道町1丁目5932番地", phone="025-266-8131",
         website="http://www.seiryo-high.ed.jp/"),
    dict(name="新潟清心女子高等学校", course=["全日制"], postal_code="950-2101",
         address="新潟市西区五十嵐1の町6370番地", phone="025-269-2041",
         website="http://www.seishin.ed.jp/"),
    dict(name="敬和学園高等学校", course=["全日制"], postal_code="950-3112",
         address="新潟市北区太夫浜325番地", phone="025-259-2391",
         website="http://www.keiwa-h.jp/"),
    dict(name="新潟第一高等学校", course=["全日制"], postal_code="951-8141",
         address="新潟市中央区関新3丁目3番1号", phone="025-231-5643",
         website="http://www.n-daiichi.ed.jp/"),
    dict(name="東京学館新潟高等学校", course=["全日制"], postal_code="950-1141",
         address="新潟市中央区鐘木185番地1", phone="025-283-8857",
         website="https://tgn.ed.jp/"),
    dict(name="日本文理高等学校", course=["全日制"], postal_code="950-2035",
         address="新潟市西区新通1072番地", phone="025-260-1000",
         website="http://www.nihonbunri.ed.jp/"),
    dict(name="帝京長岡高等学校", course=["全日制"], postal_code="940-0044",
         address="長岡市住吉3丁目9番1号", phone="0258-36-4800",
         website="http://www.teikyo-nagaoka.ed.jp/"),
    dict(name="中越高等学校", course=["全日制"], postal_code="940-8585",
         address="長岡市新保町1371番地1", phone="0258-24-0203",
         website="http://www.chuetsu-h.ed.jp/"),
    dict(name="加茂暁星高等学校", course=["全日制", "通信制"], postal_code="959-1322",
         address="加茂市学校町16番18号", phone="0256-52-2000",
         website="http://www.gyosei.ac.jp/"),
    dict(name="新発田中央高等学校", course=["全日制"], postal_code="957-8533",
         address="新発田市大字曽根570", phone="0254-27-2466",
         website="http://www.shibatachuo-h.ed.jp/"),
    dict(name="開志国際高等学校", course=["全日制"], postal_code="959-2637",
         address="胎内市長橋上439番地1", phone="0254-44-3330",
         website="http://kaishi-kokusai.ed.jp/"),
    dict(name="新潟産業大学附属高等学校", course=["全日制", "通信制"], postal_code="945-1397",
         address="柏崎市大字安田2510番地2", phone="0257-24-6644",
         website="http://www.nsf-h.ed.jp/"),
    dict(name="上越高等学校", course=["全日制"], postal_code="943-0892",
         address="上越市寺町3丁目4番34号", phone="025-523-2601",
         website="http://www.joetsu-hs.com/"),
    dict(name="関根学園高等学校", course=["全日制"], postal_code="943-8561",
         address="上越市大貫2丁目9番1号", phone="025-523-2702",
         website="http://www.sekinegakuen.com/"),
    dict(name="開志学園高等学校", course=["通信制"], postal_code="950-0931",
         address="新潟市中央区南長潟21番1号", phone="025-287-3390",
         website="http://www.kaishi.ed.jp/index.html"),
    dict(name="創進学園高等学校", course=["通信制"], postal_code="954-0051",
         address="見附市本所2丁目2番21号", phone="0258-62-0703",
         website="http://soushin.ed.jp/"),
    dict(name="新潟英智高等学校", course=["通信制"], postal_code="940-1154",
         address="長岡市宮栄3-16-14", phone="0258-31-6771",
         website="http://www.eichi.ed.jp/"),
    dict(name="開志創造高等学校", course=["通信制"], postal_code="950-0916",
         address="新潟市中央区米山3丁目1番53号", phone="025-250-0662",
         website="https://kaishi-souzou.ed.jp/"),
]

PRIVATE_MIDDLE_SCHOOLS: list[dict[str, Any]] = [
    dict(name="新潟第一中学校", postal_code="951-8141",
         address="新潟市中央区関新3丁目3番1号", phone="025-231-5643",
         website="http://www.n-daiichi.ed.jp/"),
    dict(name="新潟清心女子中学校", postal_code="950-2101",
         address="新潟市西区五十嵐1の町6370番地", phone="025-269-2041",
         website="http://www.seishin.ed.jp/"),
    dict(name="新潟明訓中学校", postal_code="950-0116",
         address="新潟市江南区北山1037番地", phone="025-257-2131",
         website="http://www.niigata-meikun.ed.jp/"),
]

PRIVATE_KINDERGARTENS: list[dict[str, Any]] = [
    dict(name="二葉幼稚園", postal_code="951-8118", address="新潟市中央区西中町714番地",
         phone="025-222-4509", website="http://www.futabakg.com/"),
    dict(name="あさひ幼稚園", postal_code="951-8122", address="新潟市中央区旭町通2番町5239",
         phone="025-222-9186", website="http://www.asahi1953.com/"),
    dict(name="あそびの森金鵄幼稚園", postal_code="950-2012", address="新潟市西区小針台1-4",
         phone="025-267-4121", website="http://kinshi-ariake.jp/"),
    dict(name="まるみ幼稚園", postal_code="950-0963", address="新潟市中央区南出来島1-17-1",
         phone="025-283-4856", website="http://www.marumiyoutien.jp/"),
    dict(name="恵光学園第一幼稚園", postal_code="950-0905", address="新潟市中央区天神尾1丁目4番1号",
         phone="025-244-2451", website="http://kei-kou.jp/"),
    dict(name="新潟青陵幼稚園", postal_code="951-8121", address="新潟市中央区水道町1丁目5939番地",
         phone="025-266-8674", website="http://www.n-seiryo.ac.jp/yochien/index.html"),
    dict(name="あおい幼稚園", postal_code="950-0801", address="新潟市東区津島屋3-100",
         phone="025-275-0772", website="https://aoi-yochien.com/"),
    dict(name="坂井輪幼稚園", postal_code="950-2041", address="新潟市西区坂井東4丁目12番8号",
         phone="025-268-2441", website="http://sakaiwa.ed.jp/"),
    dict(name="神宮幼稚園", postal_code="951-8104", address="新潟市中央区西大畑町5195番地",
         phone="025-222-7484", website="https://jingu-kindergarten.net/"),
    dict(name="聖ラファエル幼稚園", postal_code="950-0086", address="新潟市中央区花園2丁目6番7号",
         phone="025-241-2307", website="http://www.st-rafael.jp/"),
    dict(name="真人幼稚園", postal_code="950-0915", address="新潟市中央区鐙西2丁目21番17号",
         phone="025-244-4297", website="http://www.kidslink.jp/shinjin-kindergarten/"),
    dict(name="みどり幼稚園", postal_code="951-8116", address="新潟市中央区東中通1-86",
         phone="025-201-6664", website="https://midori-k.com/"),
    dict(name="鵬幼稚園", postal_code="940-0867", address="長岡市豊2丁目13-24",
         phone="0258-36-5600", website="http://ootori.ac.jp/"),
    dict(name="もみじ幼稚園", postal_code="943-0834", address="上越市西城町3丁目9番17号",
         phone="025-523-2463", website="http://www.momiji-youchien.jp/"),
    dict(name="真行寺幼稚園", postal_code="942-0001", address="上越市中央5丁目1番1号",
         phone="025-543-2829", website="http://www.nrs.ed.jp/"),
    dict(name="上越カトリック天使幼稚園", postal_code="943-0834", address="上越市西城町2丁目4番4号",
         phone="025-523-5071", website="http://www.j-tenshi.ed.jp/"),
    dict(name="認定こども園三条白百合幼稚園", postal_code="955-0046", address="三条市興野1丁目5番18号",
         phone="0256-33-1359", website="http://www.sanjo-shirayuri.com/"),
    dict(name="インマヌエル・ルーテル幼稚園", postal_code="955-0046", address="三条市興野1丁目4番15号",
         phone="0256-32-3651", website="http://www.ginzado.ne.jp/~lutheran/"),
    dict(name="小鳩幼稚園", postal_code="945-0051", address="柏崎市東本町2丁目4番5号",
         phone="0257-22-3492", website="https://www.kisnet.or.jp/~kobato/"),
    dict(name="花ぞの幼稚園", postal_code="945-0062", address="柏崎市新花町12番20号",
         phone="0257-22-4618", website="http://park11.wakwak.com/~hanazono/"),
    dict(name="柏崎二葉幼稚園", postal_code="945-0114", address="柏崎市藤井字山崎756-3",
         phone="0257-24-9050", website="http://www.kashiwazaki-futaba.com/"),
    dict(name="柏崎カトリック白百合幼稚園", postal_code="945-0831", address="柏崎市柳橋町3番10号",
         phone="0257-23-6906", website=""),
    dict(name="柏崎中央幼稚園", postal_code="945-0066", address="柏崎市西本町1丁目10番41号",
         phone="0257-23-6366", website="http://www.chuyo-tanpopo.jp/"),
    dict(name="加茂テモテ・ルーテル幼稚園", postal_code="959-1354", address="加茂市新町2丁目11番17号",
         phone="0256-52-0348", website="http://park19.wakwak.com/~ruteru/"),
    dict(name="加茂葵幼稚園", postal_code="959-1373", address="加茂市松坂町6-3",
         phone="0256-52-1984", website="http://www.kindergarten-aoi.com/"),
    dict(name="十日町カトリック天使幼稚園", postal_code="948-0051", address="十日町市寿町4丁目4-7番地",
         phone="025-752-3466", website="https://tokamachi-catholic.jp/"),
    dict(name="見附天使幼稚園", postal_code="954-0052", address="見附市学校町2丁目14番4号",
         phone="0258-62-1911", website="http://www.mitsuketenshi.jp/"),
    dict(name="つぐみ幼稚園", postal_code="954-0111", address="見附市今町5-38-18",
         phone="0258-66-6488", website="http://www.tugumi-net.com/"),
    dict(name="村上幼稚園", postal_code="958-0834", address="村上市新町2-23",
         phone="0254-52-4947", website="http://murakami-kids.com/index.html"),
    dict(name="糸魚川カトリック天使幼稚園", postal_code="941-0062", address="糸魚川市中央2丁目1番40号",
         phone="025-552-1216", website="http://park19.wakwak.com/~itoigawatenshi/"),
    dict(name="早通みずほ幼稚園", postal_code="950-3372", address="新潟市北区早通78",
         phone="025-386-9474", website="http://hayadorimizuho.com"),
    dict(name="亀田カトリック幼稚園", postal_code="950-0153", address="新潟市江南区船戸山4丁目5番7号",
         phone="025-382-7766", website="http://kameda-c-kindergarten.com/"),
    dict(name="ひのまる幼稚園", postal_code="953-0041", address="新潟市西蒲区巻甲755番地3",
         phone="0256-72-5740", website="https://www.hinomaruyouchien.jp/"),
    dict(name="和光幼稚園", postal_code="959-0425", address="新潟市西蒲区押付1361-2",
         phone="0256-88-3520", website="http://www.wako-youchien.com/"),
    dict(name="田上いずみルーテル幼稚園", postal_code="959-1502", address="南蒲原郡田上町田上丙2821-3",
         phone="0256-57-2723", website="http://www.ginzado.ne.jp/~tagami-lutheran/index.html"),
    dict(name="めぐみ幼稚園", postal_code="946-0076", address="魚沼市井口新田360",
         phone="025-792-6768", website=""),
]


def build_private_schools() -> list[dict[str, Any]]:
    records = []
    for item in PRIVATE_HIGH_SCHOOLS:
        records.append(make_record(
            name=item["name"], postal_code=item["postal_code"], address=item["address"],
            school_type="高等学校", establishment="私立", operator=item["name"],
            phone=item["phone"], website=item["website"],
            source_name="新潟県私立中学・高等学校の一覧（令和8年度）",
            source_url=PRIVATE_HS_SOURCE_URL, source_date=PRIVATE_SOURCE_DATE,
            course=item["course"],
        ))
    for item in PRIVATE_MIDDLE_SCHOOLS:
        records.append(make_record(
            name=item["name"], postal_code=item["postal_code"], address=item["address"],
            school_type="中学校", establishment="私立", operator=item["name"],
            phone=item["phone"], website=item["website"],
            source_name="新潟県私立中学・高等学校の一覧（令和8年度）",
            source_url=PRIVATE_HS_SOURCE_URL, source_date=PRIVATE_SOURCE_DATE,
        ))
    for item in PRIVATE_KINDERGARTENS:
        records.append(make_record(
            name=item["name"], postal_code=item["postal_code"], address=item["address"],
            school_type="幼稚園", establishment="私立", operator=item["name"],
            phone=item["phone"], website=item["website"],
            source_name="新潟県私立幼稚園の一覧（令和8年度、郵便番号は日本郵便データより補完）",
            source_url=PRIVATE_KG_SOURCE_URL, source_date=PRIVATE_SOURCE_DATE,
        ))
    return records


# ---------------------------------------------------------------------------
# メイン変換処理
# ---------------------------------------------------------------------------

SRC_ROOT = Path("data-source/niigata/2025")


def deduplicate(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id: dict[str, dict[str, Any]] = {}
    for record in records:
        record_id = record["id"]
        if record_id not in by_id:
            by_id[record_id] = record
            continue
        suffix = 2
        while f"{record_id}-{suffix}" in by_id:
            suffix += 1
        record["id"] = f"{record_id}-{suffix}"
        by_id[record["id"]] = record
    return list(by_id.values())


def sort_key(record: dict[str, Any]):
    m_idx = MUNICIPALITY_ORDER.index(record["municipality"]) if record["municipality"] in MUNICIPALITY_ORDER else 999
    t_idx = SCHOOL_TYPE_ORDER.index(record["school_type"]) if record["school_type"] in SCHOOL_TYPE_ORDER else 999
    return (m_idx, t_idx, record["name"])


def main() -> int:
    global SRC_ROOT
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, default=Path("data-source/niigata/2025"))
    parser.add_argument("--output", type=Path, default=Path("data/school-database/niigata.json"))
    parser.add_argument("--warnings-output", type=Path,
                         default=Path("tools/school-database/niigata_conversion_warnings.json"))
    args = parser.parse_args()
    SRC_ROOT = args.source_root

    all_records: list[dict[str, Any]] = []

    def extend(label: str, records: list[dict[str, Any]]) -> None:
        print(f"{label}: {len(records)} records")
        all_records.extend(records)

    extend("elementary.xlsx", read_pref_simple_excel(
        args.source_root / "elementary.xlsx", "小学校（全体）", school_type="小学校",
        source_name="新潟県市町村立小学校一覧（新潟市立を除く、令和8年度）", has_region_columns=True))
    extend("middle.xlsx", read_pref_simple_excel(
        args.source_root / "middle.xlsx", "中学校（全体）", school_type="中学校",
        source_name="新潟県市町村立中学校一覧（新潟市立を除く、令和8年度）", has_region_columns=True))
    extend("compulsory.xlsx", read_pref_simple_excel(
        args.source_root / "compulsory.xlsx", "義務教育学校", school_type="義務教育学校",
        source_name="新潟県市町村立義務教育学校一覧（新潟市立を除く、令和8年度）", has_region_columns=False))
    extend("pref_middle_secondary.xlsx", read_pref_simple_excel(
        args.source_root / "pref_middle_secondary.xlsx", "県立中・中等", school_type="中等教育学校",
        source_name="新潟県立中学校・中等教育学校一覧（令和8年度）", has_region_columns=False))
    extend("special_needs.xlsx", read_pref_special_needs_excel(args.source_root / "special_needs.xlsx"))

    extend("niigatacity_elementary.xlsx", read_city_two_col_excel(
        args.source_root / "niigatacity_elementary.xlsx", "小学校", school_type="小学校",
        source_name="令和8年度新潟市立学校一覧 小学校の部"))
    extend("niigatacity_middle.xlsx", read_city_two_col_excel(
        args.source_root / "niigatacity_middle.xlsx", "中学校", school_type="中学校",
        source_name="令和8年度新潟市立学校一覧 中学校の部"))
    extend("niigatacity_kindergarten.xlsx", read_city_kindergarten_excel(
        args.source_root / "niigatacity_kindergarten.xlsx"))
    extend("niigatacity small schools (high/secondary/special-needs)",
           build_niigata_city_small_schools())

    extend("gakkou_gairan.txt (pref high/secondary schools)", build_pref_high_schools())

    extend("private schools (high/middle/kindergarten)", build_private_schools())

    all_records = deduplicate(all_records)
    all_records.sort(key=sort_key)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(all_records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"TOTAL: {len(all_records)} records -> {args.output}")

    args.warnings_output.parent.mkdir(parents=True, exist_ok=True)
    args.warnings_output.write_text(
        json.dumps(WARNINGS.items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
