#!/usr/bin/env python3
"""神奈川県公式Excel名簿を学校検索用JSONへ変換する。

原本は data-source/kanagawa/2025/ に置く（Git管理外）。

公立は学校種・設置者ごとに個別ファイルに分かれている（17ファイル、基準日は
すべて令和8年4月1日）。私立は1ブック7シート（幼稚園/小学校/中学校/
義務教育学校/高等学校/中等教育学校/特別支援学校、専修学校・各種学校は対象外）
にまとまっている。千葉県版と異なり、新設・統合等の別途補正資料は公式サイトに
見当たらなかったため、原本の基準日がそのまま最新の状態を反映しているものとして
補正処理は行わない。

学校種の扱い:
  - 高等学校・中等教育学校・特別支援学校は原則1レコード1行（原本の時点で
    課程等が統合済み）。高等学校の「課程」列に複数値（例:「全日・定時単位」）が
    入っている場合は course 配列へ分割する。
  - 義務教育学校は前期課程・後期課程で所在地が2行に分かれる学校（横浜市立
    霧が丘義務教育学校）があるが、原本の行構成をそのまま保持し、統合は行わない
    （課程ごとに所在地が異なるため）。
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# 正規化ユーティリティ（千葉県版と共通の設計）
# ---------------------------------------------------------------------------

def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    text = text.replace("\n", "").replace("\r", "")
    return text.strip()


def normalize_name(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"[ \t　]+", "", text)


def normalize_postal_code(value: Any) -> str:
    if value is None:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 7:
        return f"{digits[:3]}-{digits[3:]}"
    return normalize_text(value).replace("〒", "")


PAREN_AREA_CODE_RE = re.compile(r"^\((\d+)\)(\d+-\d+)$")


def normalize_phone(value: Any) -> str:
    if value is None:
        return ""
    text = normalize_text(value)
    if text in ("", "-", "―", "ー"):
        return ""
    text = text.replace("−", "-").replace("―", "-").replace("ー", "-")
    match = PAREN_AREA_CODE_RE.match(text)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    if "-" in text:
        return text
    digits = re.sub(r"\D", "", text)
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    return text


def normalize_address(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("−", "-").replace("―", "-")
    if text and not text.startswith("神奈川県"):
        text = "神奈川県" + text
    return text


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).lower()
    normalized = re.sub(r"[^0-9a-zぁ-んァ-ヶ一-龠]+", "-", normalized).strip("-")
    return normalized or "school"


# 「横浜市\n（本校333校\n分校2校）」のような集計注記を取り除く。
# この注記は「平塚市\n（本園１校）」のように市町村名と同一セルに同居する場合と、
# 行が縦に長い注記（本校333校/分校2校）ではセル自体が複数行にまたがり、
# 丸括弧が開き/閉じで別々のセル（別の物理行）に分かれてしまう場合の両方がある。
# そのため (1) まず対になった丸括弧の注記を除去し、(2) 残った文字列に数字や
# 本校/分校等のキーワード、片方だけの丸括弧が残っていれば注記の残骸とみなして
# 破棄する（=Noneを返しforward-fill側で直前の値を維持させる）。
ANNOTATION_RE = re.compile(r"[（(][^（）()]*[）)]")
ANNOTATION_KEYWORDS = ("本校", "分校", "本園", "分園", "分教室")


def clean_carry_cell(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    text = ANNOTATION_RE.sub("", text).strip()
    if not text:
        return None
    if re.search(r"\d", text) or any(ch in text for ch in "（）()"):
        return None
    if any(keyword in text for keyword in ANNOTATION_KEYWORDS):
        return None
    return text


# ---------------------------------------------------------------------------
# 市町村表示順（横浜市各区 -> 川崎市各区 -> 相模原市各区 -> 市部 -> 郡部）
# ---------------------------------------------------------------------------

YOKOHAMA_WARDS = [
    "横浜市鶴見区", "横浜市神奈川区", "横浜市西区", "横浜市中区", "横浜市南区",
    "横浜市港南区", "横浜市保土ケ谷区", "横浜市旭区", "横浜市磯子区", "横浜市金沢区",
    "横浜市港北区", "横浜市緑区", "横浜市青葉区", "横浜市都筑区", "横浜市戸塚区",
    "横浜市栄区", "横浜市泉区", "横浜市瀬谷区",
]
KAWASAKI_WARDS = [
    "川崎市川崎区", "川崎市幸区", "川崎市中原区", "川崎市高津区",
    "川崎市宮前区", "川崎市多摩区", "川崎市麻生区",
]
SAGAMIHARA_WARDS = ["相模原市緑区", "相模原市中央区", "相模原市南区"]

# 市部（横須賀三浦 -> 県央 -> 湘南 -> 県西、県公式「地域別市町村一覧」の掲載順に準拠）
KANAGAWA_CITIES = [
    "横須賀市", "鎌倉市", "逗子市", "三浦市",
    "厚木市", "大和市", "海老名市", "座間市", "綾瀬市",
    "平塚市", "藤沢市", "茅ヶ崎市", "秦野市", "伊勢原市",
    "小田原市", "南足柄市",
]

# 郡部（三浦郡 -> 高座郡 -> 中郡 -> 愛甲郡 -> 足柄上郡 -> 足柄下郡）
KANAGAWA_GUN_TOWNS = [
    "三浦郡葉山町",
    "高座郡寒川町",
    "中郡大磯町", "中郡二宮町",
    "愛甲郡愛川町", "愛甲郡清川村",
    "足柄上郡中井町", "足柄上郡大井町", "足柄上郡松田町", "足柄上郡山北町", "足柄上郡開成町",
    "足柄下郡箱根町", "足柄下郡真鶴町", "足柄下郡湯河原町",
]

MUNICIPALITY_ORDER = (
    YOKOHAMA_WARDS + KAWASAKI_WARDS + SAGAMIHARA_WARDS + KANAGAWA_CITIES + KANAGAWA_GUN_TOWNS
)

_BARE_TOWN_TO_CANONICAL = {
    re.match(r"(三浦郡|高座郡|中郡|愛甲郡|足柄上郡|足柄下郡)(.+)", t).group(2): t
    for t in KANAGAWA_GUN_TOWNS
}
def _match_form(value: str) -> str:
    """「保土ケ谷区」表記ゆれ（ヶ/ケ/ヵ/カ）を吸収するための照合専用の正規化。
    戻り値の municipality 自体は MUNICIPALITY_ORDER の正式表記を使うため、
    住所文字列そのものは書き換えない。"""
    return value.replace("ヶ", "ケ").replace("ヵ", "カ")


_MUNICIPALITY_CANDIDATES = sorted(
    ((_match_form(c), c) for c in MUNICIPALITY_ORDER), key=lambda pair: len(pair[0]), reverse=True
)
_BARE_TOWN_CANDIDATES = sorted(
    ((_match_form(bare), bare) for bare in _BARE_TOWN_TO_CANONICAL), key=lambda pair: len(pair[0]), reverse=True
)


def infer_municipality(address: str) -> str:
    text = address
    if text.startswith("神奈川県"):
        text = text[len("神奈川県"):]
    text_for_match = _match_form(text)
    for match_candidate, canonical in _MUNICIPALITY_CANDIDATES:
        if text_for_match.startswith(match_candidate):
            return canonical
    for match_bare, bare in _BARE_TOWN_CANDIDATES:
        if text_for_match.startswith(match_bare):
            return _BARE_TOWN_TO_CANONICAL[bare]
    return ""


SCHOOL_TYPE_ORDER = [
    "幼稚園", "小学校", "中学校", "義務教育学校",
    "高等学校", "中等教育学校", "特別支援学校",
]


def is_suspended_notice(value: Any) -> bool:
    """電話番号欄に「※現在休園中」「※休校中」等の休止注記が入っている行を検出する。
    埼玉県版の運用（休校中の学校は公開対象から除外する）にならい、
    このような学校は原本に掲載されていても収録対象から除外する。"""
    text = normalize_text(value)
    return any(keyword in text for keyword in ("休園", "休校", "休止", "休部"))


class WarningLog:
    def __init__(self) -> None:
        self.items: list[dict[str, str]] = []

    def add(self, context: str, message: str) -> None:
        self.items.append({"context": context, "message": message})


WARNINGS = WarningLog()


def make_record(
    *, name: str, name_kana: str, postal_code: str, address: str, school_type: str,
    establishment: str, operator: str, phone: str, website: str,
    source_name: str, source_url: str, source_date: str, course: list[str],
) -> dict[str, Any]:
    name = normalize_name(name)
    address = normalize_address(address)
    municipality = infer_municipality(address)
    stable_key = "|".join((establishment, school_type, municipality, name, ",".join(course)))
    return {
        "id": f"kanagawa-{slug(stable_key)}",
        "prefecture": "神奈川県",
        "name": name,
        "name_kana": normalize_name(name_kana),
        "postal_code": normalize_postal_code(postal_code),
        "address": address,
        "municipality": municipality,
        "school_type": school_type,
        "establishment": establishment,
        "operator": normalize_name(operator),
        "phone": normalize_phone(phone),
        "website": website,
        "source_name": source_name,
        "source_url": source_url,
        "source_date": source_date,
        "verified_date": "",
        "course": list(course),
    }


def split_course(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [c for c in re.split(r"[・\s　]+", text) if c]


# ---------------------------------------------------------------------------
# 公立ファイル読み込み
# ---------------------------------------------------------------------------

PUBLIC_SOURCE_DATE = "2026-04-01"
PUBLIC_SOURCE_BASE = "https://www.pref.kanagawa.jp/docs/u5t/gakkoumeibo/index.html"


def read_public_file(
    path: Path, *, school_type: str, has_ward: bool, establishment_fixed: str | None,
    course_col: int | None, name_col: int, kana_col: int, postal_col: int,
    address_col: int, phone_col: int, muni_col: int = 2, ward_col: int | None = None,
    source_name: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))

    records: list[dict[str, Any]] = []
    current_muni = ""
    current_ward = ""

    for row in rows[4:]:  # ヘッダーは4行目まで
        def cell(col: int) -> Any:
            idx = col - 1
            return row[idx] if idx < len(row) else None

        muni_candidate = clean_carry_cell(cell(muni_col))
        if muni_candidate:
            current_muni = muni_candidate
        if has_ward and ward_col:
            ward_candidate = clean_carry_cell(cell(ward_col))
            if ward_candidate:
                current_ward = ward_candidate

        name = normalize_name(cell(name_col))
        if not name:
            continue
        if any(word in name for word in ("計", "合計", "学校数")):
            continue

        address = normalize_address(cell(address_col))
        if not address:
            continue

        if is_suspended_notice(cell(phone_col)):
            WARNINGS.add("suspended", f"休園・休校中のため除外: {name}（{path.name}）")
            continue

        if establishment_fixed:
            establishment = establishment_fixed
        else:
            establishment = "国立" if "国立" in current_muni else "公立"

        operator = (current_muni + current_ward) if has_ward else current_muni

        course = split_course(cell(course_col)) if course_col else []

        record = make_record(
            name=name,
            name_kana=cell(kana_col),
            postal_code=cell(postal_col),
            address=address,
            school_type=school_type,
            establishment=establishment,
            operator=operator,
            phone=cell(phone_col),
            website="",
            source_name=source_name,
            source_url=f"{PUBLIC_SOURCE_BASE}#{path.name}",
            source_date=PUBLIC_SOURCE_DATE,
            course=course,
        )
        records.append(record)
    return records


PUBLIC_FILE_SPECS = [
    dict(file="01_r8_preschool.xlsx", school_type="幼稚園", has_ward=False,
         establishment_fixed="公立", course_col=None, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 幼稚園（市立・町立・村立、令和8年4月1日現在）"),
    dict(file="02_r8_primaryschool_yokohama.xlsx", school_type="小学校", has_ward=True,
         establishment_fixed="公立", course_col=None, muni_col=2, ward_col=3, name_col=4,
         kana_col=5, postal_col=6, address_col=7, phone_col=8,
         source_name="神奈川県公立学校名簿 小学校（横浜市、令和8年4月1日現在）"),
    dict(file="03_r8_primaryschool_kawasaki.xlsx", school_type="小学校", has_ward=True,
         establishment_fixed="公立", course_col=None, muni_col=2, ward_col=3, name_col=4,
         kana_col=5, postal_col=6, address_col=7, phone_col=8,
         source_name="神奈川県公立学校名簿 小学校（川崎市、令和8年4月1日現在）"),
    dict(file="04_r8_primaryschool_sagamihara.xlsx", school_type="小学校", has_ward=True,
         establishment_fixed="公立", course_col=None, muni_col=2, ward_col=3, name_col=4,
         kana_col=5, postal_col=6, address_col=7, phone_col=8,
         source_name="神奈川県公立学校名簿 小学校（相模原市、令和8年4月1日現在）"),
    dict(file="05_r8_primaryschool.xlsx", school_type="小学校", has_ward=False,
         establishment_fixed="公立", course_col=None, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 小学校（横浜市・川崎市・相模原市以外、令和8年4月1日現在）"),
    dict(file="06_r8_primaryschool_national.xlsx", school_type="小学校", has_ward=False,
         establishment_fixed="国立", course_col=None, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 小学校（国立大学法人、令和8年4月1日現在）"),
    dict(file="07_r8_juniorhighschool_yokohama.xlsx", school_type="中学校", has_ward=True,
         establishment_fixed="公立", course_col=None, muni_col=2, ward_col=3, name_col=4,
         kana_col=5, postal_col=6, address_col=7, phone_col=8,
         source_name="神奈川県公立学校名簿 中学校（横浜市、令和8年4月1日現在）"),
    dict(file="08_r8_juniorhighschool_kawasaki.xlsx", school_type="中学校", has_ward=True,
         establishment_fixed="公立", course_col=None, muni_col=2, ward_col=3, name_col=4,
         kana_col=5, postal_col=6, address_col=7, phone_col=8,
         source_name="神奈川県公立学校名簿 中学校（川崎市、令和8年4月1日現在）"),
    dict(file="09_r8_juniorhighschool_sagamihara.xlsx", school_type="中学校", has_ward=True,
         establishment_fixed="公立", course_col=None, muni_col=2, ward_col=3, name_col=4,
         kana_col=5, postal_col=6, address_col=7, phone_col=8,
         source_name="神奈川県公立学校名簿 中学校（相模原市、令和8年4月1日現在）"),
    dict(file="10_r8_juniorhighschool.xlsx", school_type="中学校", has_ward=False,
         establishment_fixed="公立", course_col=None, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 中学校（横浜市・川崎市・相模原市以外、令和8年4月1日現在）"),
    dict(file="11_r8_juniorhighschool_national.xlsx", school_type="中学校", has_ward=False,
         establishment_fixed="国立", course_col=None, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 中学校（国立大学法人、令和8年4月1日現在）"),
    dict(file="12_r8_highschool_pref.xlsx", school_type="高等学校", has_ward=False,
         establishment_fixed="公立", course_col=9, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 高等学校（県立、令和8年4月1日現在）"),
    dict(file="13_r8_highschool_city.xlsx", school_type="高等学校", has_ward=False,
         establishment_fixed="公立", course_col=9, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 高等学校（市立、令和8年4月1日現在）"),
    dict(file="14_r8_compulsoryeducationschool.xlsx", school_type="義務教育学校", has_ward=True,
         establishment_fixed="公立", course_col=None, muni_col=2, ward_col=3, name_col=4,
         kana_col=5, postal_col=6, address_col=7, phone_col=8,
         source_name="神奈川県公立学校名簿 義務教育学校（横浜市・相模原市、令和8年4月1日現在）"),
    dict(file="15_r8_six-yearsecondaryschool.xlsx", school_type="中等教育学校", has_ward=False,
         establishment_fixed="公立", course_col=None, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 中等教育学校（県立、令和8年4月1日現在）"),
    dict(file="16_r8_special-needsschool.xlsx", school_type="特別支援学校", has_ward=False,
         establishment_fixed=None, course_col=None, muni_col=2, name_col=3,
         kana_col=4, postal_col=5, address_col=6, phone_col=7,
         source_name="神奈川県公立学校名簿 特別支援学校（県立・市立・国立大学法人、令和8年4月1日現在）"),
]


# ---------------------------------------------------------------------------
# 私立ファイル読み込み（1ブック7シート）
# ---------------------------------------------------------------------------

PRIVATE_SOURCE_NAME = "神奈川県私立学校名簿（令和8年5月1日現在）"
PRIVATE_SOURCE_URL = "https://www.pref.kanagawa.jp/documents/22470/r80501~meibo.xlsx"
PRIVATE_SOURCE_DATE = "2026-05-01"

PRIVATE_SHEET_SPECS = [
    dict(sheet="幼稚園", school_type="幼稚園", has_number=True, has_muni=True,
         name_col=3, postal_col=4, address_col=5, phone_col=6, operator_col=7, course_col=None),
    dict(sheet="小学校", school_type="小学校", has_number=False, has_muni=False,
         name_col=1, postal_col=2, address_col=3, phone_col=4, operator_col=5, course_col=None),
    dict(sheet="中学校", school_type="中学校", has_number=True, has_muni=False,
         name_col=2, postal_col=3, address_col=4, phone_col=5, operator_col=6, course_col=None),
    dict(sheet="義務教育学校", school_type="義務教育学校", has_number=True, has_muni=False,
         name_col=2, postal_col=3, address_col=4, phone_col=5, operator_col=6, course_col=None),
    dict(sheet="高等学校", school_type="高等学校", has_number=True, has_muni=False,
         name_col=2, postal_col=4, address_col=5, phone_col=6, operator_col=7, course_col=3),
    dict(sheet="中等教育学校", school_type="中等教育学校", has_number=False, has_muni=False,
         name_col=1, postal_col=2, address_col=3, phone_col=4, operator_col=5, course_col=None),
    dict(sheet="特別支援学校", school_type="特別支援学校", has_number=False, has_muni=False,
         name_col=1, postal_col=2, address_col=3, phone_col=4, operator_col=5, course_col=None),
]


def read_private_workbook(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    all_records: list[dict[str, Any]] = []

    for spec in PRIVATE_SHEET_SPECS:
        if spec["sheet"] not in workbook.sheetnames:
            WARNINGS.add("private", f"シート未検出: {spec['sheet']}")
            continue
        sheet = workbook[spec["sheet"]]
        rows = list(sheet.iter_rows(values_only=True))

        for row in rows[1:]:  # ヘッダーは1行のみ（2物理行にまたがるが1論理行）
            def cell(col: int) -> Any:
                idx = col - 1
                return row[idx] if idx < len(row) else None

            name = normalize_name(cell(spec["name_col"]))
            if not name:
                continue

            address = normalize_address(cell(spec["address_col"]))
            if not address:
                continue

            if is_suspended_notice(cell(spec["phone_col"])):
                WARNINGS.add("suspended", f"休園・休校中のため除外: {name}（私立名簿 {spec['sheet']}）")
                continue

            course = split_course(cell(spec["course_col"])) if spec["course_col"] else []

            record = make_record(
                name=name,
                name_kana="",
                postal_code=cell(spec["postal_col"]),
                address=address,
                school_type=spec["school_type"],
                establishment="私立",
                operator=cell(spec["operator_col"]),
                phone=cell(spec["phone_col"]),
                website="",
                source_name=PRIVATE_SOURCE_NAME,
                source_url=PRIVATE_SOURCE_URL,
                source_date=PRIVATE_SOURCE_DATE,
                course=course,
            )
            all_records.append(record)
    return all_records


# ---------------------------------------------------------------------------
# メイン変換処理
# ---------------------------------------------------------------------------





def deduplicate_exact(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged = {}
    ordered_keys = []
    
    for r in records:
        name_val = r.get("school_name", r.get("name", ""))
        est_val = r.get("establishment", r.get("establishment_type", ""))
        
        sig = (
            r.get("prefecture", ""),
            r.get("municipality", ""),
            name_val,
            r.get("school_type", ""),
            est_val,
            r.get("postal_code", ""),
            r.get("address", ""),
            r.get("phone", ""),
            r.get("operator", "")
        )
        
        if sig not in merged:
            merged[sig] = r.copy()
            if "course" not in merged[sig] or merged[sig]["course"] is None:
                pass
            elif isinstance(merged[sig]["course"], str):
                merged[sig]["course"] = [merged[sig]["course"]]
            elif not isinstance(merged[sig]["course"], list):
                merged[sig]["course"] = list(merged[sig]["course"])
            ordered_keys.append(sig)
        else:
            current_courses = merged[sig].get("course", [])
            if not isinstance(current_courses, list):
                current_courses = [current_courses] if current_courses else []
                
            new_courses = r.get("course", [])
            if not isinstance(new_courses, list):
                new_courses = [new_courses] if new_courses else []
                
            for c in new_courses:
                if c and c not in current_courses:
                    current_courses.append(c)
            
            if "course" in merged[sig] or current_courses:
                merged[sig]["course"] = current_courses
            
    return [merged[k] for k in ordered_keys]


def deduplicate(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_id: dict[str, dict[str, Any]] = {}
    for record in records:
        record_id = record["id"]
        if record_id not in by_id:
            by_id[record_id] = record
            continue
        suffix = 2
        while f"{record_id}-{suffix}" in by_id:
            suffix += 1
        record["id"] = f"{record_id}-{suffix}"
        by_id[record["id"]] = record
    return list(by_id.values())


def sort_key(record: dict[str, Any]):
    m_idx = MUNICIPALITY_ORDER.index(record["municipality"]) if record["municipality"] in MUNICIPALITY_ORDER else 999
    t_idx = SCHOOL_TYPE_ORDER.index(record["school_type"]) if record["school_type"] in SCHOOL_TYPE_ORDER else 999
    return (m_idx, t_idx, record["name"])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", type=Path, default=Path("data-source/kanagawa/2025"))
    parser.add_argument("--output", type=Path, default=Path("data/school-database/kanagawa.json"))
    parser.add_argument("--warnings-output", type=Path,
                         default=Path("tools/school-database/kanagawa_conversion_warnings.json"))
    args = parser.parse_args()

    all_records: list[dict[str, Any]] = []

    for spec in PUBLIC_FILE_SPECS:
        path = args.source_root / spec["file"]
        if not path.exists():
            WARNINGS.add("public", f"SKIP missing: {path}")
            continue
        records = read_public_file(
            path,
            school_type=spec["school_type"],
            has_ward=spec["has_ward"],
            establishment_fixed=spec["establishment_fixed"],
            course_col=spec["course_col"],
            muni_col=spec["muni_col"],
            ward_col=spec.get("ward_col"),
            name_col=spec["name_col"],
            kana_col=spec["kana_col"],
            postal_col=spec["postal_col"],
            address_col=spec["address_col"],
            phone_col=spec["phone_col"],
            source_name=spec["source_name"],
        )
        print(f"{spec['file']}: {len(records)} records")
        all_records.extend(records)

    private_path = args.source_root / "r80501_shiritsu_meibo.xlsx"
    if private_path.exists():
        private_records = read_private_workbook(private_path)
        print(f"{private_path.name}: {len(private_records)} records")
        all_records.extend(private_records)
    else:
        WARNINGS.add("private", f"SKIP missing: {private_path}")

    all_records = deduplicate_exact(all_records)
    all_records = deduplicate(all_records)
    all_records.sort(key=sort_key)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(all_records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\nWrote {len(all_records)} records to {args.output}")

    args.warnings_output.parent.mkdir(parents=True, exist_ok=True)
    args.warnings_output.write_text(
        json.dumps(WARNINGS.items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {len(WARNINGS.items)} warnings to {args.warnings_output}")
    for item in WARNINGS.items:
        print(f"WARN [{item['context']}] {item['message']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
